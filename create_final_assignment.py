# -*- coding: utf-8 -*-
"""
분리 규칙을 적용한 학생 배정 및 엑셀 생성 (4반/3반 편성 선택)

- 4반: 1,2,3,4반 → A,B,C,D 자유 배정 (기존과 동일)
- 3반: 1→B,C,D / 2→A,C,D / 3→A,B,D / 4→A,B,C (선생님 연속 지도 배제)
환경변수 ASSIGN_MODE로 "4반" 또는 "3반" 지정.
"""
import pandas as pd
import random
import time
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from class_assignment_logic import (
    calculate_class_distribution,
    calculate_transfer_plan
)
from class_assignment_logic import calculate_transfer_plan_3반, ALLOWED_TARGETS_3반

# 랜덤 시드 설정
# - 기본값: 매 실행마다 달라지도록 현재 시간 기반 시드 사용
# - 재현이 필요하면 환경변수 ASSIGN_SEED에 정수를 지정 (예: 12345)
_seed_env = os.environ.get("ASSIGN_SEED")
try:
    BASE_SEED = int(_seed_env) if _seed_env is not None else int(time.time())
except Exception:
    BASE_SEED = int(time.time())
random.seed(BASE_SEED)
print(f"[SEED] BASE_SEED={BASE_SEED} (재현 필요 시 PowerShell에서: $env:ASSIGN_SEED={BASE_SEED})")

ASSIGN_MODE = os.environ.get("ASSIGN_MODE", "4반")
print(f"[MODE] ASSIGN_MODE={ASSIGN_MODE} (4반: 자유 배정 / 3반: 선생님 연속 지도 배제)")

# 데이터 읽기
df_students = pd.read_excel('학생자료.xlsx', sheet_name=0)
df_separation = pd.read_excel('separation.xlsx', sheet_name=0)

# 학번 첫자리에서 현재 학년 추출
current_grade = 4  # 기본값
for _, row in df_students.iterrows():
    student_id = str(row['학번'])
    if student_id and len(student_id) > 0 and student_id[0].isdigit():
        current_grade = int(student_id[0])
        break
print(f"[학년] 학번 첫자리에서 추출한 현재 학년: {current_grade}학년")

# 분리 규칙 딕셔너리 생성 및 빈도수 계산
separation_graph = {}
all_separation_ids = []

for _, row in df_separation.iterrows():
    id1 = row['학생1']
    id2 = row['학생2']
    all_separation_ids.extend([id1, id2])
    
    if id1 not in separation_graph:
        separation_graph[id1] = []
    if id2 not in separation_graph:
        separation_graph[id2] = []
    separation_graph[id1].append(id2)
    separation_graph[id2].append(id1)

# 빈도수 계산
frequency = Counter(all_separation_ids)

# separation.xlsx에 등장한 학생만 필터링
separation_student_ids = set(separation_graph.keys())
df_separation_students = df_students[df_students['학번'].isin(separation_student_ids)].copy()

# separation.xlsx에 없는 학생들 (수동 배정용)
df_manual_students = df_students[~df_students['학번'].isin(separation_student_ids)].copy()

print("=" * 80)
print("분리 규칙 분석")
print("=" * 80)
print(f"분리 규칙에 등장한 학생 수: {len(separation_student_ids)}명")
print(f"수동 배정 필요한 학생 수: {len(df_manual_students)}명")
print(f"\n빈도수 상위 10명:")
for student_id, freq in frequency.most_common(10):
    student = df_students[df_students['학번'] == student_id].iloc[0]
    print(f"  {student['이름']} ({student_id}): {freq}회")
print()

# 전체 현황 (전체 학생자료.xlsx 기준)
total_students = len(df_students)
male_count = len(df_students[df_students['남녀'] == '남'])
female_count = len(df_students[df_students['남녀'] == '여'])

# 각 이전학반의 현황 (전체 학생자료.xlsx 기준)
previous_class_counts = {}
for prev_class in [1, 2, 3, 4]:
    prev_df = df_students[df_students['이전학반'] == prev_class]
    previous_class_counts[prev_class] = {
        'male': len(prev_df[prev_df['남녀'] == '남']),
        'female': len(prev_df[prev_df['남녀'] == '여'])
    }

# 목표 반 배정 계산 (전체 학생 기준)
target_distribution = calculate_class_distribution(total_students, male_count, female_count)
if ASSIGN_MODE == "3반":
    transfer_plan = calculate_transfer_plan_3반(previous_class_counts, target_distribution)
else:
    transfer_plan = calculate_transfer_plan(previous_class_counts, target_distribution)

# 디버깅: 전체 목표 분배 확인
print("=" * 80)
print("전체 목표 분배")
print("=" * 80)
print(f"총 학생 수: {total_students}명 (남 {male_count}명, 여 {female_count}명)")
print(f"A반: 총 {target_distribution['total']['A']}명 (남 {target_distribution['male']['A']}명, 여 {target_distribution['female']['A']}명)")
print(f"B반: 총 {target_distribution['total']['B']}명 (남 {target_distribution['male']['B']}명, 여 {target_distribution['female']['B']}명)")
print(f"C반: 총 {target_distribution['total']['C']}명 (남 {target_distribution['male']['C']}명, 여 {target_distribution['female']['C']}명)")
print(f"D반: 총 {target_distribution['total']['D']}명 (남 {target_distribution['male']['D']}명, 여 {target_distribution['female']['D']}명)")
print()

# 디버깅: transfer_plan 확인
print("=" * 80)
print("transfer_plan 확인")
print("=" * 80)
print(f"[중요] 현재 ASSIGN_MODE: {ASSIGN_MODE}")
for prev_class in [1, 2, 3, 4]:
    male_vals = [transfer_plan[prev_class]['male'][t] for t in ['A','B','C','D']]
    female_vals = [transfer_plan[prev_class]['female'][t] for t in ['A','B','C','D']]
    print(f"{prev_class}반 남학생: A={male_vals[0]}, B={male_vals[1]}, C={male_vals[2]}, D={male_vals[3]} (합계: {sum(male_vals)}명)")
    print(f"{prev_class}반 여학생: A={female_vals[0]}, B={female_vals[1]}, C={female_vals[2]}, D={female_vals[3]} (합계: {sum(female_vals)}명)")
    # 4반 편성 모드에서 이전학반 4반이 A로 배정되지 않으면 경고
    if ASSIGN_MODE == "4반" and prev_class == 4:
        if male_vals[0] == 0 and female_vals[0] == 0:
            print(f"  [경고] 4반 편성 모드인데 이전학반 4반이 A반으로 배정되지 않았습니다!")
        else:
            print(f"  [확인] 이전학반 4반이 A반으로 배정됨 (남 {male_vals[0]}명, 여 {female_vals[0]}명)")
print()

# 디버깅: 실제로 모이는 학생 수 확인
print("=" * 80)
print("실제로 모이는 학생 수 확인")
print("=" * 80)
for target in ['A', 'B', 'C', 'D']:
    actual_male = sum(transfer_plan[prev]['male'][target] for prev in [1, 2, 3, 4])
    actual_female = sum(transfer_plan[prev]['female'][target] for prev in [1, 2, 3, 4])
    actual_total = actual_male + actual_female
    target_male = target_distribution['male'][target]
    target_female = target_distribution['female'][target]
    target_total = target_distribution['total'][target]
    print(f"{target}반: 실제 {actual_total}명 (남 {actual_male}명, 여 {actual_female}명) / 목표 {target_total}명 (남 {target_male}명, 여 {target_female}명)")
print()

target_classes = ['A', 'B', 'C', 'D']
previous_classes = [1, 2, 3, 4]

# 각 반에 배정될 학생 리스트 초기화
class_assignments = {target: {'male': [], 'female': []} for target in target_classes}

# 배정 점수 계산 함수 (각 이전학반 내에서 A,B,C,D 균일 배정 고려)
def calculate_assignment_score(student, target, class_assignments, transfer_plan, prev_class, target_distribution, df_students, gender):
    """
    학생을 특정 반에 배정할 때의 점수 계산
    점수가 낮을수록 좋은 배정
    각 이전학반 내에서 A,B,C,D가 균일하게 배정되도록 고려
    남학생과 여학생을 각각 독립적으로 균형 맞춤
    """
    score = 0
    target_classes = ['A', 'B', 'C', 'D']
    
    # gender를 transfer_plan의 키 형식으로 변환 ('남'/'여' -> 'male'/'female')
    gender_key = 'male' if gender == '남' else 'female'
    
    # 1. 각 이전학반 내에서 해당 성별의 A,B,C,D 배정 수 균형 체크 (가장 중요)
    # 현재 이전학반에서 각 반에 배정된 학생 수
    prev_class_counts = {}
    for t in target_classes:
        if gender == '남':
            prev_class_counts[t] = len([s for s in class_assignments[t]['male'] if s['이전학반'] == prev_class])
        else:
            prev_class_counts[t] = len([s for s in class_assignments[t]['female'] if s['이전학반'] == prev_class])
    
    # 배정 후 예상 값
    prev_class_counts_after = prev_class_counts.copy()
    prev_class_counts_after[target] += 1
    
    # 이전학반 내에서 최대값과 최소값의 차이 (균형 지표)
    before_imbalance = max(prev_class_counts.values()) - min(prev_class_counts.values())
    after_imbalance = max(prev_class_counts_after.values()) - min(prev_class_counts_after.values())
    
    # 균형 개선도 (차이가 줄어들수록 좋음)
    balance_improvement = before_imbalance - after_imbalance
    # 가중치를 매우 높게 설정하여 이전학반 내 균형을 최우선으로 고려
    score -= balance_improvement * 1000
    
    # 2. 배정 가능한 공간이 많은 반부터 먼저 채우기
    # needed: 목표 인원수, current: 현재 배정된 인원수
    needed = transfer_plan[prev_class][gender_key][target]
    current = prev_class_counts[target]
    remaining_space = needed - current  # 배정 가능한 공간
    
    # 남은 공간이 많을수록 우선 배정 (점수 낮춤)
    score -= remaining_space * 100
    
    # 3. 전체 반별 균일 배정 (보조적 고려)
    current_total = len(class_assignments[target]['male']) + len(class_assignments[target]['female'])
    target_total = target_distribution['total'][target]
    # 목표 인원수와의 차이 (차이가 작을수록 좋음)
    score += abs(current_total - target_total) * 1  # 가중치 낮게
    
    return score

# 자리바꾸기 함수 (학력수준/제2외국어/영어반 균등 배분)
def try_swap_students(class_assignments, separation_graph, df_students, target_classes):
    """
    두 학생을 교환하여 학력수준/제2외국어/영어반 분포를 더 균등하게 만듦
    분리 규칙을 위반하지 않는 경우만 교환
    """
    improved = False
    
    # 모든 반 쌍에 대해 교환 시도
    for i, target1 in enumerate(target_classes):
        for target2 in target_classes[i+1:]:
            # 각 반의 모든 학생 쌍에 대해 교환 시도
            students1 = class_assignments[target1]['male'] + class_assignments[target1]['female']
            students2 = class_assignments[target2]['male'] + class_assignments[target2]['female']
            
            for student1 in students1:
                for student2 in students2:
                    id1 = student1['학번']
                    id2 = student2['학번']
                    
                    # 같은 이전학반이 아니면 교환 불가 (transfer_plan 제약)
                    if student1['이전학반'] != student2['이전학반']:
                        continue
                    
                    # 같은 성별이 아니면 교환 불가
                    if student1['남녀'] != student2['남녀']:
                        continue
                    
                    # 분리 규칙 확인: 교환 후에도 위반하지 않는지 확인
                    can_swap = True
                    
                    # student1이 target2로 가는 경우
                    if id1 in separation_graph:
                        for separated_id in separation_graph[id1]:
                            for s in students2:
                                if s['학번'] == separated_id and s['학번'] != id2:
                                    can_swap = False
                                    break
                            if not can_swap:
                                break
                    
                    # student2가 target1로 가는 경우
                    if can_swap and id2 in separation_graph:
                        for separated_id in separation_graph[id2]:
                            for s in students1:
                                if s['학번'] == separated_id and s['학번'] != id1:
                                    can_swap = False
                                    break
                            if not can_swap:
                                break
                    
                    if not can_swap:
                        continue
                    
                    # 교환 후 분포 개선 여부 확인
                    improvement = calculate_swap_improvement(
                        student1, student2, target1, target2, 
                        class_assignments, df_students
                    )
                    
                    if improvement > 0:
                        # 교환 실행
                        if student1['남녀'] == '남':
                            class_assignments[target1]['male'].remove(student1)
                            class_assignments[target2]['male'].remove(student2)
                            class_assignments[target1]['male'].append(student2)
                            class_assignments[target2]['male'].append(student1)
                        else:
                            class_assignments[target1]['female'].remove(student1)
                            class_assignments[target2]['female'].remove(student2)
                            class_assignments[target1]['female'].append(student2)
                            class_assignments[target2]['female'].append(student1)
                        
                        improved = True
                        break
                
                if improved:
                    break
            
            if improved:
                break
        
        if improved:
            break
    
    return improved

def calculate_swap_improvement(student1, student2, target1, target2, class_assignments, df_students):
    """
    두 학생을 교환했을 때의 분포 개선 정도 계산
    양수면 개선, 음수면 악화
    """
    improvement = 0
    
    # 학생 정보 가져오기
    info1 = df_students[df_students['학번'] == student1['학번']]
    info2 = df_students[df_students['학번'] == student2['학번']]
    
    if len(info1) == 0 or len(info2) == 0:
        return 0
    
    row1 = info1.iloc[0]
    row2 = info2.iloc[0]
    
    # 각 반의 현재 분포 (교환 전)
    ids1 = [s['학번'] for s in class_assignments[target1]['male'] + class_assignments[target1]['female']]
    ids2 = [s['학번'] for s in class_assignments[target2]['male'] + class_assignments[target2]['female']]
    
    students1 = df_students[df_students['학번'].isin(ids1)]
    students2 = df_students[df_students['학번'].isin(ids2)]
    
    # 학력수준 분포 개선
    if '학력수준' in row1 and '학력수준' in row2 and pd.notna(row1['학력수준']) and pd.notna(row2['학력수준']):
        level1 = row1['학력수준']
        level2 = row2['학력수준']
        
        if level1 != level2:
            # 교환 전: target1에 level1, target2에 level2
            count1_level1 = len(students1[students1['학력수준'] == level1])
            count2_level2 = len(students2[students2['학력수준'] == level2])
            count1_level2 = len(students1[students1['학력수준'] == level2])
            count2_level1 = len(students2[students2['학력수준'] == level1])
            
            # 교환 후: target1에 level2 증가, level1 감소, target2에 level1 증가, level2 감소
            # 분포가 더 균등해지면 개선
            before_imbalance = abs(count1_level1 - count2_level1) + abs(count1_level2 - count2_level2)
            after_imbalance = abs((count1_level1 - 1) - (count2_level1 + 1)) + abs((count1_level2 + 1) - (count2_level2 - 1))
            improvement += (before_imbalance - after_imbalance) * 0.1
    
    # 제2외국어 분포 개선
    if '제2외국어' in row1 and '제2외국어' in row2 and pd.notna(row1['제2외국어']) and pd.notna(row2['제2외국어']):
        lang1 = row1['제2외국어']
        lang2 = row2['제2외국어']
        
        if lang1 != lang2:
            count1_lang1 = len(students1[students1['제2외국어'] == lang1])
            count2_lang2 = len(students2[students2['제2외국어'] == lang2])
            count1_lang2 = len(students1[students1['제2외국어'] == lang2])
            count2_lang1 = len(students2[students2['제2외국어'] == lang1])
            
            before_imbalance = abs(count1_lang1 - count2_lang1) + abs(count1_lang2 - count2_lang2)
            after_imbalance = abs((count1_lang1 - 1) - (count2_lang1 + 1)) + abs((count1_lang2 + 1) - (count2_lang2 - 1))
            improvement += (before_imbalance - after_imbalance) * 0.1
    
    # 영어반 분포 개선
    if '영어반' in row1 and '영어반' in row2 and pd.notna(row1['영어반']) and pd.notna(row2['영어반']):
        eng1 = row1['영어반']
        eng2 = row2['영어반']
        
        if eng1 != eng2:
            count1_eng1 = len(students1[students1['영어반'] == eng1])
            count2_eng2 = len(students2[students2['영어반'] == eng2])
            count1_eng2 = len(students1[students1['영어반'] == eng2])
            count2_eng1 = len(students2[students2['영어반'] == eng1])
            
            before_imbalance = abs(count1_eng1 - count2_eng1) + abs(count1_eng2 - count2_eng2)
            after_imbalance = abs((count1_eng1 - 1) - (count2_eng1 + 1)) + abs((count1_eng2 + 1) - (count2_eng2 - 1))
            improvement += (before_imbalance - after_imbalance) * 0.1
    
    return improvement

def get_available_targets(student_id, prev_class, gender, class_assignments, transfer_plan, 
                          separation_graph, target_classes, allowed_targets_map):
    """
    특정 학생이 배정 가능한 반 목록을 반환 (분리규칙 + transfer_plan 고려)
    """
    targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes
    available = []
    
    for target in targets_for_prev:
        # transfer_plan 확인
        needed = transfer_plan[prev_class][gender][target]
        current = len([s for s in class_assignments[target][gender] if s['이전학반'] == prev_class])
        if current >= needed:
            continue
        
        # 분리 규칙 확인
        can_assign = True
        if student_id in separation_graph:
            for separated_id in separation_graph[student_id]:
                for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                    if assigned_student['학번'] == separated_id:
                        can_assign = False
                        break
                if not can_assign:
                    break
        
        if can_assign:
            available.append(target)
    
    return available

def assign_students_with_retry(class_assignments, df_separation_students, separation_graph, transfer_plan, 
                                target_distribution, df_students, previous_classes, target_classes, 
                                frequency, separation_student_ids, random_seed=None, allowed_targets_map=None):
    """
    학생 배정을 시도하고, 분리 규칙 위반 수를 반환
    우선순위 기반 배정: 빈도 높은 순 → 실패 시 실패 학생 최우선으로 재배정
    allowed_targets_map: 3반 편성 시 이전학반별 배정 가능 반 (None이면 4반 자유 배정)
    """
    if random_seed is not None:
        random.seed(random_seed)
    
    # 배정 초기화
    for target in target_classes:
        class_assignments[target]['male'] = []
        class_assignments[target]['female'] = []
    
    # 모든 학생을 한 리스트에 수집 (separation.xlsx에 등장한 학생만)
    all_students_to_assign = []
    for prev_class in previous_classes:
        prev_df_all = df_students[df_students['이전학반'] == prev_class].copy()
        prev_df = prev_df_all[prev_df_all['학번'].isin(separation_student_ids)].copy()
        
        for _, row in prev_df.iterrows():
            all_students_to_assign.append({
                '학번': row['학번'],
                '이름': row['이름'],
                '남녀': row['남녀'],
                '이전학반': row['이전학반']
            })
    
    # 1차: 빈도(분리대상) 높은 순으로 정렬
    all_students_to_assign.sort(key=lambda s: (-frequency.get(s['학번'], 0), random.random()))
    
    # 반복 배정: 실패한 학생을 맨 앞으로 이동하면서 시도
    max_retry = 20  # 최대 20번 재정렬 시도
    for retry_count in range(max_retry):
        failed_students = []
        temp_assignments = {t: {'male': [], 'female': []} for t in target_classes}
        
        # 현재 순서대로 배정 시도
        for student in all_students_to_assign:
            student_id = student['학번']
            prev_class = student['이전학반']
            gender = 'male' if student['남녀'] == '남' else 'female'
            
            # 배정 가능한 반 목록
            available = get_available_targets(
                student_id, prev_class, gender, temp_assignments, 
                transfer_plan, separation_graph, target_classes, allowed_targets_map
            )
            
            if not available:
                # 배정 실패 → 실패 목록에 추가
                failed_students.append(student)
                continue
            
            # 가용한 반 중에서 점수가 가장 좋은 반 선택
            best_target = None
            best_score = float('inf')
            
            for target in available:
                score = calculate_assignment_score(
                    student, target, temp_assignments, transfer_plan, 
                    prev_class, target_distribution, df_students, student['남녀']
                )
                if score < best_score:
                    best_score = score
                    best_target = target
            
            # 배정 실행
            if best_target:
                temp_assignments[best_target][gender].append(student)
            else:
                failed_students.append(student)
        
        # 실패한 학생이 없으면 성공
        if not failed_students:
            # 임시 배정을 실제 배정에 복사
            for target in target_classes:
                class_assignments[target]['male'] = temp_assignments[target]['male'][:]
                class_assignments[target]['female'] = temp_assignments[target]['female'][:]
            break
        
        # 실패한 학생이 있으면 순서 재조정: 실패 학생 맨 앞 + 나머지
        successfully_assigned = [s for s in all_students_to_assign if s not in failed_students]
        all_students_to_assign = failed_students + successfully_assigned
        
        # 재시도 시 약간의 랜덤성 추가
        random.shuffle(failed_students)
        all_students_to_assign = failed_students + successfully_assigned
    
    # 배정되지 않은 학생 수 계산
    assigned_ids = set()
    for target in target_classes:
        for student in class_assignments[target]['male'] + class_assignments[target]['female']:
            assigned_ids.add(student['학번'])
    
    all_ids = set(s['학번'] for s in all_students_to_assign)
    unassigned_count = len(all_ids - assigned_ids)
    
    if unassigned_count > 0:
        return float('inf'), unassigned_count
    
    # 이전 로직의 남은 부분 제거 (아래 코드는 더 이상 필요 없음)
    # 하지만 기존 로직과의 호환성을 위해 남은 학생 재배정 로직은 유지하지 않음
    
    # 미배정 학생 수 계산 (이미 all_students_to_assign이 비어있으므로 0)
    unassigned_count = 0
    
    # 이전 코드에서 계속되던 남학생/여학생 개별 배정 로직을 제거하고
    # 바로 분리 규칙 위반 확인으로 이동
    
    # 남은 학생 재배정 로직도 제거 (MRV가 이미 최적화했으므로)
    
    # 분리 규칙 위반 확인으로 바로 이동
    # 아래는 기존 코드 제거를 위한 placeholder
    
    # 디버깅: 1반 남학생 배정 결과 확인 (선택적)
    # 생략...
    
    # 분리 규칙 위반 확인
    violations = []
    violation_pairs = {}
    
    for target in target_classes:
        all_students = class_assignments[target]['male'] + class_assignments[target]['female']
        student_ids = [s['학번'] for s in all_students]
        
        for i, student1 in enumerate(all_students):
            id1 = student1['학번']
            if id1 in separation_graph:
                for separated_id in separation_graph[id1]:
                    if separated_id in student_ids:
                        student2 = next(s for s in all_students if s['학번'] == separated_id)
                        pair_key = tuple(sorted([id1, separated_id]))
                        
                        if pair_key not in violation_pairs:
                            violation_pairs[pair_key] = {
                                '반': target,
                                '학생1': {'학번': id1, '이름': student1['이름']},
                                '학생2': {'학번': separated_id, '이름': student2['이름']}
                            }
                            violations.append(violation_pairs[pair_key])
    
    # 미배정 학생이 있거나 분리 규칙 위반이 있으면 반환
    return len(violations), unassigned_count

# ======== 기존 코드 제거: 아래부터는 실행되지 않음 ========
if False:
    # 각 반별로 분리 규칙을 만족하는지 확인하고 점수 계산
    candidates = []
    
    for target in targets_for_prev:
                needed = male_transfers[target]
                current = len([s for s in class_assignments[target]['male'] if s['이전학반'] == prev_class])
                
                # transfer_plan을 초과하지 않도록
                if current >= needed:
                    continue
                
                # 분리 규칙 확인
                can_assign = True
                if student_id in separation_graph:
                    for separated_id in separation_graph[student_id]:
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if not can_assign:
                    continue
                
                # 분리 규칙을 만족하는 경우만 후보에 추가
                score = calculate_assignment_score(student, target, class_assignments, transfer_plan, prev_class, target_distribution, df_students, '남')
                candidates.append((score, target, 0))
            
            # 분리 규칙을 만족하는 반에만 배정 (위반 시 배정하지 않음)
            best_target = None
            if candidates:
                candidates.sort(key=lambda x: x[0])
                best_target = candidates[0][1]
                class_assignments[best_target]['male'].append(student)
            # 만족하는 반이 없으면 배정하지 않음 → remaining으로 수집됨
        
        # 여학생 배정 - separation.xlsx에 등장한 학생들만 배정
        # transfer_plan에 정해진 인원수에 맞춰 가능한 한 균등하게 배정
        female_transfers = transfer_plan[prev_class]['female']
        targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes

        # 각 학생에 대해 배정 가능한 반을 찾고, transfer_plan에 맞춰 배정
        for student in prev_female:
            student_id = student['학번']
            
            already_assigned = False
            for t in target_classes:
                if any(s['학번'] == student_id for s in class_assignments[t]['male'] + class_assignments[t]['female']):
                    already_assigned = True
                    break
            
            if already_assigned:
                continue
            
            # 각 반별로 분리 규칙을 만족하는지 확인하고 점수 계산
            candidates = []
            
            for target in targets_for_prev:
                needed = female_transfers[target]
                current = len([s for s in class_assignments[target]['female'] if s['이전학반'] == prev_class])
                
                # transfer_plan을 초과하지 않도록
                if current >= needed:
                    continue
                
                # 분리 규칙 확인
                can_assign = True
                if student_id in separation_graph:
                    for separated_id in separation_graph[student_id]:
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if not can_assign:
                    continue
                
                # 분리 규칙을 만족하는 경우만 후보에 추가
                score = calculate_assignment_score(student, target, class_assignments, transfer_plan, prev_class, target_distribution, df_students, '여')
                candidates.append((score, target, 0))
            
            # 분리 규칙을 만족하는 반에만 배정 (위반 시 배정하지 않음)
            best_target = None
            if candidates:
                candidates.sort(key=lambda x: x[0])
                best_target = candidates[0][1]
                class_assignments[best_target]['female'].append(student)
            # 만족하는 반이 없으면 배정하지 않음 → remaining으로 수집됨
        
        # 디버깅: 1반 남학생 배정 결과 확인
        if prev_class == 1:
            print(f"\n{prev_class}반 남학생 배정 결과:")
            for target in target_classes:
                count = len([s for s in class_assignments[target]['male'] if s['이전학반'] == prev_class])
                print(f"  {target}반: {count}명 (목표: {transfer_plan[prev_class]['male'][target]}명)")
            print(f"{prev_class}반 여학생 배정 결과:")
            for target in target_classes:
                count = len([s for s in class_assignments[target]['female'] if s['이전학반'] == prev_class])
                print(f"  {target}반: {count}명 (목표: {transfer_plan[prev_class]['female'][target]}명)")
    
    # 남은 학생들 배정 (separation.xlsx에 등장한 학생 중 배정되지 않은 학생)
    remaining_male = []
    remaining_female = []
    
    for prev_class in previous_classes:
        # separation.xlsx에 등장한 학생 중에서
        prev_df = df_separation_students[df_separation_students['이전학반'] == prev_class].copy()
        prev_male = prev_df[prev_df['남녀'] == '남'].to_dict('records')
        prev_female = prev_df[prev_df['남녀'] == '여'].to_dict('records')
        
        for student in prev_male:
            student_id = student['학번']
            assigned = False
            for t in target_classes:
                if any(s['학번'] == student_id for s in class_assignments[t]['male'] + class_assignments[t]['female']):
                    assigned = True
                    break
            if not assigned:
                remaining_male.append(student)
        
        for student in prev_female:
            student_id = student['학번']
            assigned = False
            for t in target_classes:
                if any(s['학번'] == student_id for s in class_assignments[t]['male'] + class_assignments[t]['female']):
                    assigned = True
                    break
            if not assigned:
                remaining_female.append(student)
    
    # 남은 학생들을 여러 번 반복하여 최적 배정 시도
    max_remaining_iterations = 30  # 최대 30번 반복 (미배정 학생 해결을 위해 증가)
    for iteration in range(max_remaining_iterations):
        remaining_male.sort(key=lambda s: (-frequency.get(s['학번'], 0), random.random()))
        remaining_female.sort(key=lambda s: (-frequency.get(s['학번'], 0), random.random()))
        
        new_remaining_male = []
        new_remaining_female = []
        
        for student in remaining_male:
            student_id = student['학번']
            prev_class = student['이전학반']
            targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes

            best_target = None
            best_score = float('inf')
            
            for target in targets_for_prev:
                needed = transfer_plan[prev_class]['male'][target]
                current = len([s for s in class_assignments[target]['male'] if s['이전학반'] == prev_class])
                
                can_assign = True
                if student_id in separation_graph:
                    for separated_id in separation_graph[student_id]:
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if not can_assign:
                    continue
                
                score = calculate_assignment_score(student, target, class_assignments, transfer_plan, prev_class, target_distribution, df_students, '남')
                if current >= needed:
                    score += (current - needed + 1) * 500
                
                if score < best_score:
                    best_score = score
                    best_target = target
            
            if best_target:
                class_assignments[best_target]['male'].append(student)
            else:
                new_remaining_male.append(student)
        
        for student in remaining_female:
            student_id = student['학번']
            prev_class = student['이전학반']
            targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes

            best_target = None
            best_score = float('inf')
            
            for target in targets_for_prev:
                needed = transfer_plan[prev_class]['female'][target]
                current = len([s for s in class_assignments[target]['female'] if s['이전학반'] == prev_class])
                
                can_assign = True
                if student_id in separation_graph:
                    for separated_id in separation_graph[student_id]:
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if not can_assign:
                    continue
                
                score = calculate_assignment_score(student, target, class_assignments, transfer_plan, prev_class, target_distribution, df_students, '여')
                if current >= needed:
                    score += (current - needed + 1) * 500
                
                if score < best_score:
                    best_score = score
                    best_target = target
            
            if best_target:
                class_assignments[best_target]['female'].append(student)
            else:
                new_remaining_female.append(student)
        
        remaining_male = new_remaining_male
        remaining_female = new_remaining_female
        
        # 더 이상 배정할 수 없으면 중단
        if not remaining_male and not remaining_female:
            break
    
    # 미배정 학생 수 계산
    unassigned_count = len(remaining_male) + len(remaining_female)
    
    # 미배정 학생 디버깅 정보 출력 (첫 시도에서만)
    if unassigned_count > 0 and random_seed == BASE_SEED:
        print(f"\n[디버깅] 미배정 학생 {unassigned_count}명:")
        for student in remaining_male[:5]:
            student_id = student['학번']
            prev_class = student['이전학반']
            sep_count = len(separation_graph.get(student_id, []))
            targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes
            print(f"  - {student['이름']} ({student_id}, 이전학반 {prev_class}반): 분리대상 {sep_count}명, 배정가능반 {targets_for_prev}")
            # 각 반에 분리대상이 몇 명 있는지 확인
            if student_id in separation_graph:
                for target in targets_for_prev:
                    assigned_in_target = [s['학번'] for s in class_assignments[target]['male'] + class_assignments[target]['female']]
                    blocked_by = [sid for sid in separation_graph[student_id] if sid in assigned_in_target]
                    if blocked_by:
                        print(f"    {target}반: 분리대상 {len(blocked_by)}명 배정됨 (배정 불가)")
        for student in remaining_female[:5]:
            student_id = student['학번']
            prev_class = student['이전학반']
            sep_count = len(separation_graph.get(student_id, []))
            targets_for_prev = allowed_targets_map[prev_class] if allowed_targets_map else target_classes
            print(f"  - {student['이름']} ({student_id}, 이전학반 {prev_class}반): 분리대상 {sep_count}명, 배정가능반 {targets_for_prev}")
            # 각 반에 분리대상이 몇 명 있는지 확인
            if student_id in separation_graph:
                for target in targets_for_prev:
                    assigned_in_target = [s['학번'] for s in class_assignments[target]['male'] + class_assignments[target]['female']]
                    blocked_by = [sid for sid in separation_graph[student_id] if sid in assigned_in_target]
                    if blocked_by:
                        print(f"    {target}반: 분리대상 {len(blocked_by)}명 배정됨 (배정 불가)")
    
    # 분리 규칙 위반 확인
    violations = []
    violation_pairs = {}
    
    for target in target_classes:
        all_students = class_assignments[target]['male'] + class_assignments[target]['female']
        student_ids = [s['학번'] for s in all_students]
        
        for i, student1 in enumerate(all_students):
            id1 = student1['학번']
            if id1 in separation_graph:
                for separated_id in separation_graph[id1]:
                    if separated_id in student_ids:
                        student2 = next(s for s in all_students if s['학번'] == separated_id)
                        pair_key = tuple(sorted([id1, separated_id]))
                        
                        if pair_key not in violation_pairs:
                            violation_pairs[pair_key] = {
                                '반': target,
                                '학생1': {'학번': id1, '이름': student1['이름']},
                                '학생2': {'학번': separated_id, '이름': student2['이름']}
                            }
                            violations.append(violation_pairs[pair_key])
    
    # 미배정 학생이 있으면 (위반수, 미배정수) 반환, 없으면 (위반수, 0) 반환
    return len(violations), unassigned_count

# 여러 번 시도하여 분리 규칙 위반이 0개인 배정 찾기
print("=" * 80)
print("학생 배정 시도 중 (우선순위 기반 재배정)")
print("1단계: 빈도(분리대상) 높은 순으로 배정")
print("2단계: 실패 학생을 최우선으로 이동 후 재배정 반복")
print("=" * 80)

best_assignments = None
best_violation_count = float('inf')
best_unassigned_count = float('inf')
max_attempts = 500  # 최대 500번 시도 (MRV 휴리스틱 적용으로 더 많은 시도 가능)

for attempt in range(max_attempts):
    # 배정 초기화
    class_assignments = {
        'A': {'male': [], 'female': []},
        'B': {'male': [], 'female': []},
        'C': {'male': [], 'female': []},
        'D': {'male': [], 'female': []}
    }
    
    violation_count, unassigned_count = assign_students_with_retry(
        class_assignments, df_separation_students, separation_graph, transfer_plan,
        target_distribution, df_students, previous_classes, target_classes,
        frequency, separation_student_ids, random_seed=(BASE_SEED + attempt),
        allowed_targets_map=ALLOWED_TARGETS_3반 if ASSIGN_MODE == "3반" else None
    )
    
    # 미배정 학생이 없고 위반도 적은 배정을 우선
    # 미배정 학생이 있으면 큰 페널티 부여 (미배정 1명 = 위반 1000개로 취급)
    effective_score = violation_count + (unassigned_count * 1000)
    current_best_score = best_violation_count + (best_unassigned_count * 1000) if best_assignments else float('inf')
    
    if effective_score < current_best_score:
        best_violation_count = violation_count
        best_unassigned_count = unassigned_count
        # 딥카피로 저장
        import copy
        best_assignments = copy.deepcopy(class_assignments)
        
        if unassigned_count > 0:
            print(f"시도 {attempt + 1}: 분리 규칙 위반 {violation_count}개, 미배정 {unassigned_count}명 (최선 기록)")
        else:
            print(f"시도 {attempt + 1}: 분리 규칙 위반 {violation_count}개, 모든 학생 배정 완료 (최선 기록)")
        
        # 위반이 0개이고 미배정도 없으면 즉시 중단
        if violation_count == 0 and unassigned_count == 0:
            print(f"완벽한 배정을 찾았습니다! (시도 {attempt + 1}회)")
            break
    elif attempt % 20 == 0:
        if unassigned_count > 0:
            print(f"시도 {attempt + 1}: 위반 {violation_count}개, 미배정 {unassigned_count}명 (최선: 위반 {best_violation_count}개, 미배정 {best_unassigned_count}명)")
        else:
            print(f"시도 {attempt + 1}: 위반 {violation_count}개 (최선: 위반 {best_violation_count}개)")

# 미배정 학생이 있으면 경고 출력
if best_unassigned_count > 0:
    print(f"\n[경고] 최선의 배정에서도 {best_unassigned_count}명의 학생이 배정되지 않았습니다.")
    print("분리규칙을 모두 만족하면서 배정할 수 없는 상태입니다.")

# 최선의 배정 사용
if best_assignments:
    class_assignments = best_assignments
else:
    # 최선의 배정이 없으면 마지막 배정 사용
    pass

print(f"\n최종 결과: 분리 규칙 위반 {best_violation_count}개")

# 분리 규칙 위반이 0개인 경우에만 자리바꾸기 실행
if best_violation_count == 0:
    print("자리바꾸기로 분포 개선 중...")
    swap_count = 0
    for swap_iteration in range(10):  # 최대 10번 반복
        improved = try_swap_students(class_assignments, separation_graph, df_students, target_classes)
        if improved:
            swap_count += 1
        else:
            break  # 더 이상 개선할 수 없으면 중단
    
    print(f"자리바꾸기 실행: {swap_count}회")
else:
    print("분리 규칙 위반이 있으므로 자리바꾸기를 실행하지 않습니다.")

# 분리 규칙 위반 확인 (최종)
violations = []
violation_pairs = {}

for target in target_classes:
    all_students = class_assignments[target]['male'] + class_assignments[target]['female']
    student_ids = [s['학번'] for s in all_students]
    
    for i, student1 in enumerate(all_students):
        id1 = student1['학번']
        if id1 in separation_graph:
            for separated_id in separation_graph[id1]:
                if separated_id in student_ids:
                    student2 = next(s for s in all_students if s['학번'] == separated_id)
                    pair_key = tuple(sorted([id1, separated_id]))
                    
                    if pair_key not in violation_pairs:
                        violation_pairs[pair_key] = {
                            '반': target,
                            '학생1': {'학번': id1, '이름': student1['이름']},
                            '학생2': {'학번': separated_id, '이름': student2['이름']}
                        }
                        violations.append(violation_pairs[pair_key])

# 위반이 불가피한지 확인
def check_if_violation_inevitable(violation, class_assignments, transfer_plan, separation_graph, target_classes):
    """
    위반이 불가피한지 확인
    두 학생을 다른 반으로 옮길 수 있는지 확인
    """
    id1 = violation['학생1']['학번']
    id2 = violation['학생2']['학번']
    current_target = violation['반']
    
    # 학생1의 이전학반과 성별 찾기
    prev_class1 = None
    gender1 = None
    for target in target_classes:
        for student in class_assignments[target]['male'] + class_assignments[target]['female']:
            if student['학번'] == id1:
                prev_class1 = student['이전학반']
                gender1 = student['남녀']
                break
        if prev_class1:
            break
    
    # 학생2의 이전학반과 성별 찾기
    prev_class2 = None
    gender2 = None
    for target in target_classes:
        for student in class_assignments[target]['male'] + class_assignments[target]['female']:
            if student['학번'] == id2:
                prev_class2 = student['이전학반']
                gender2 = student['남녀']
                break
        if prev_class2:
            break
    
    # 다른 반으로 옮길 수 있는지 확인
    can_move_student1 = False
    can_move_student2 = False
    
    for target in target_classes:
        if target == current_target:
            continue
        
        # 학생1을 옮길 수 있는지 확인
        if prev_class1:
            needed1 = transfer_plan[prev_class1]['male' if gender1 == '남' else 'female'][target]
            current1 = len([s for s in class_assignments[target]['male' if gender1 == '남' else 'female'] 
                           if s['이전학반'] == prev_class1])
            
            if current1 < needed1:
                # 분리 규칙 확인 (학생2는 제외)
                can_assign = True
                if id1 in separation_graph:
                    for separated_id in separation_graph[id1]:
                        if separated_id == id2:  # 학생2는 이미 같은 반이므로 체크 안 함
                            continue
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if can_assign:
                    can_move_student1 = True
        
        # 학생2를 옮길 수 있는지 확인
        if prev_class2:
            needed2 = transfer_plan[prev_class2]['male' if gender2 == '남' else 'female'][target]
            current2 = len([s for s in class_assignments[target]['male' if gender2 == '남' else 'female'] 
                           if s['이전학반'] == prev_class2])
            
            if current2 < needed2:
                can_assign = True
                if id2 in separation_graph:
                    for separated_id in separation_graph[id2]:
                        if separated_id == id1:  # 학생1은 이미 같은 반이므로 체크 안 함
                            continue
                        for assigned_student in class_assignments[target]['male'] + class_assignments[target]['female']:
                            if assigned_student['학번'] == separated_id:
                                can_assign = False
                                break
                        if not can_assign:
                            break
                
                if can_assign:
                    can_move_student2 = True
    
    # 둘 다 옮길 수 없으면 불가피한 위반
    return not (can_move_student1 or can_move_student2)

# 위반 그룹 생성
violation_groups = []
student_to_group = {}

for pair_key, violation in violation_pairs.items():
    id1, id2 = pair_key
    
    group1 = student_to_group.get(id1)
    group2 = student_to_group.get(id2)
    
    if group1 is None and group2 is None:
        new_group_id = len(violation_groups)
        violation_groups.append({id1, id2})
        student_to_group[id1] = new_group_id
        student_to_group[id2] = new_group_id
    elif group1 is not None and group2 is not None:
        if group1 != group2:
            violation_groups[group1].update(violation_groups[group2])
            for student_id in violation_groups[group2]:
                student_to_group[student_id] = group1
            violation_groups[group2] = None
    elif group1 is not None:
        violation_groups[group1].add(id2)
        student_to_group[id2] = group1
    else:
        violation_groups[group2].add(id1)
        student_to_group[id1] = group2

violation_groups = [g for g in violation_groups if g is not None]

# 위반 그룹별 색상
violation_colors = [
    'FF0000', '0000FF', '008000', 'FF00FF', 'FFA500',
    '800080', '008080', 'FF1493', '00CED1', 'FF4500',
]

violation_color_map = {}
for group_id, group in enumerate(violation_groups):
    color = violation_colors[group_id % len(violation_colors)]
    for student_id in group:
        violation_color_map[student_id] = color

for violation in violations:
    id1 = violation['학생1']['학번']
    violation['group_id'] = student_to_group.get(id1, -1)

print("=" * 80)
print("배정 결과 요약")
print("=" * 80)
for target in target_classes:
    actual_male = len(class_assignments[target]['male'])
    actual_female = len(class_assignments[target]['female'])
    actual_total = actual_male + actual_female
    target_male = target_distribution['male'][target]
    target_female = target_distribution['female'][target]
    target_total = target_distribution['total'][target]
    match_status = "[OK]" if (actual_male == target_male and actual_female == target_female) else "[X]"
    print(f"{target}반: 실제 {actual_total}명 (남 {actual_male}명, 여 {actual_female}명) / 목표 {target_total}명 (남 {target_male}명, 여 {target_female}명) {match_status}")

# 배정되지 않은 학생(범용) 확인
assigned_ids = set()
for target in target_classes:
    for student in class_assignments[target]['male'] + class_assignments[target]['female']:
        assigned_ids.add(student['학번'])

all_student_ids = set(df_students['학번'].tolist())
unassigned_ids = sorted(list(all_student_ids - assigned_ids))

if unassigned_ids:
    print("\n" + "!" * 80)
    print("!" * 80)
    print(f"!!!!! 경고: 배정되지 않은 학생 {len(unassigned_ids)}명 발견 !!!!!")
    print("!" * 80)
    print("!" * 80)
    print("\n미배정 학생 목록:")
    unassigned_students = df_students[df_students['학번'].isin(unassigned_ids)].to_dict('records')
    for s in unassigned_students:
        student_id = s.get('학번','')
        student_name = s.get('이름','')
        prev_class = s.get('이전학반','')
        
        # 분리대상 확인
        if student_id in separation_graph:
            sep_list = separation_graph[student_id]
            sep_names = []
            for sep_id in sep_list:
                sep_student = df_students[df_students['학번'] == sep_id]
                if not sep_student.empty:
                    sep_names.append(f"{sep_student.iloc[0]['이름']}({sep_id})")
            print(f"  - {student_name} ({student_id}, {prev_class}반) - 분리대상 {len(sep_list)}명: {', '.join(sep_names)}")
        else:
            print(f"  - {student_name} ({student_id}, {prev_class}반) - 분리대상 없음")
    
    print("\n" + "!" * 80)
    print("분리규칙을 모두 만족하면서 배정할 수 없는 상태입니다.")
    print("해결 방법:")
    print("1. 분리명부에서 일부 분리규칙을 제거")
    print("2. 프로그램을 다시 실행하여 다른 배정 시도 (랜덤 시드 변경)")
    print("!" * 80 + "\n")
else:
    print("\n배정되지 않은 학생 확인:")
    print("  [OK] 모든 학생이 배정되었습니다!")

print(f"\n분리 규칙 위반: {len(violations)}개")
if len(violations) > 0:
    print("위반한 학생 쌍:")
    for v in violations:
        print(f"  - {v['학생1']['이름']} - {v['학생2']['이름']} ({v['반']}반)")
print(f"수동 배정 필요한 학생: {len(df_manual_students)}명")

# 엑셀 파일 생성
wb = Workbook()
ws = wb.active
ws.title = "반편성 배정표"

colors = {
    'A': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
    'B': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
    'C': PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'),
    'D': PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
}

header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
header_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center')
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

current_row = 1

# 각 이전학반별로 테이블 생성
for prev_class in previous_classes:
    ws.merge_cells(f'A{current_row}:P{current_row}')
    header_cell = ws[f'A{current_row}']
    header_cell.value = f'{current_grade}학년 {prev_class}반'
    header_cell.fill = header_fill
    header_cell.font = Font(bold=True, size=12)
    header_cell.alignment = center_align
    
    ws.merge_cells(f'Q{current_row}:AF{current_row}')
    header_cell2 = ws[f'Q{current_row}']
    header_cell2.value = f'{current_grade}학년 {prev_class}반'
    header_cell2.fill = header_fill
    header_cell2.font = Font(bold=True, size=12)
    header_cell2.alignment = center_align
    current_row += 1
    
    # 헤더 행
    headers_male = ['학년', '반', '배정반', '인원수']
    for i in range(1, 7):
        headers_male.append(f'남학생{i}(학번)')
        headers_male.append(f'남학생{i}(이름)')
    
    for col_idx, header in enumerate(headers_male, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    
    headers_female = ['학년', '반', '배정반', '인원수']
    for i in range(1, 7):
        headers_female.append(f'여학생{i}(학번)')
        headers_female.append(f'여학생{i}(이름)')
    
    for col_idx, header in enumerate(headers_female, 17):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    
    current_row += 1
    
    # 각 목표 반별로 데이터 행 생성
    for target in target_classes:
        planned_male = transfer_plan[prev_class]['male'][target]
        planned_female = transfer_plan[prev_class]['female'][target]
        
        prev_male_students = [s for s in class_assignments[target]['male'] if s['이전학반'] == prev_class]
        prev_female_students = [s for s in class_assignments[target]['female'] if s['이전학반'] == prev_class]
        
        # 남학생 데이터
        ws.cell(row=current_row, column=1).value = current_grade
        ws.cell(row=current_row, column=2).value = prev_class
        ws.cell(row=current_row, column=3).value = target
        ws.cell(row=current_row, column=4).value = planned_male
        
        # transfer_plan에 정해진 인원수만큼 색칠 (배정된 학생이 없어도)
        for i in range(1, min(planned_male + 1, 7)):
            col_학번 = 4 + (i-1)*2 + 1
            col_이름 = 4 + (i-1)*2 + 2
            
            cell_학번 = ws.cell(row=current_row, column=col_학번)
            cell_이름 = ws.cell(row=current_row, column=col_이름)
            
            cell_학번.fill = colors[target]
            cell_이름.fill = colors[target]
            cell_학번.border = border_style
            cell_이름.border = border_style
            cell_학번.alignment = center_align
            cell_이름.alignment = center_align
            
            if i <= len(prev_male_students):
                student = prev_male_students[i-1]
                cell_학번.value = student['학번']
                cell_이름.value = student['이름']
                
                student_id = student['학번']
                if student_id in violation_color_map:
                    violation_color = violation_color_map[student_id]
                    cell_학번.font = Font(color=violation_color, bold=True)
                    cell_이름.font = Font(color=violation_color, bold=True)
                else:
                    cell_학번.font = Font()
                    cell_이름.font = Font()
        
        # 여학생 데이터
        ws.cell(row=current_row, column=17).value = current_grade
        ws.cell(row=current_row, column=18).value = prev_class
        ws.cell(row=current_row, column=19).value = target
        ws.cell(row=current_row, column=20).value = planned_female
        
        for i in range(1, min(planned_female + 1, 7)):
            col_학번 = 20 + (i-1)*2 + 1
            col_이름 = 20 + (i-1)*2 + 2
            
            cell_학번 = ws.cell(row=current_row, column=col_학번)
            cell_이름 = ws.cell(row=current_row, column=col_이름)
            
            cell_학번.fill = colors[target]
            cell_이름.fill = colors[target]
            cell_학번.border = border_style
            cell_이름.border = border_style
            cell_학번.alignment = center_align
            cell_이름.alignment = center_align
            
            if i <= len(prev_female_students):
                student = prev_female_students[i-1]
                cell_학번.value = student['학번']
                cell_이름.value = student['이름']
                
                student_id = student['학번']
                if student_id in violation_color_map:
                    violation_color = violation_color_map[student_id]
                    cell_학번.font = Font(color=violation_color, bold=True)
                    cell_이름.font = Font(color=violation_color, bold=True)
                else:
                    cell_학번.font = Font()
                    cell_이름.font = Font()
        
        # 테두리 적용
        for col in range(1, 17):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_style
            cell.alignment = center_align
        
        for col in range(17, 33):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_style
            cell.alignment = center_align
        
        current_row += 1
    
    # 합계 행 (색칠된 칸수 기준)
    prev_total_male = sum(transfer_plan[prev_class]['male'][t] for t in target_classes)
    prev_total_female = sum(transfer_plan[prev_class]['female'][t] for t in target_classes)
    prev_total = prev_total_male + prev_total_female
    
    ws.cell(row=current_row, column=1).value = current_grade
    ws.cell(row=current_row, column=2).value = prev_class
    ws.cell(row=current_row, column=3).value = "합계"
    ws.cell(row=current_row, column=4).value = prev_total_male
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    
    ws.cell(row=current_row, column=17).value = current_grade
    ws.cell(row=current_row, column=18).value = prev_class
    ws.cell(row=current_row, column=19).value = "합계"
    ws.cell(row=current_row, column=20).value = prev_total_female
    ws.cell(row=current_row, column=20).font = Font(bold=True)
    
    for col in range(1, 17):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = border_style
    
    for col in range(17, 33):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = border_style
    
    ws.merge_cells(f'E{current_row}:P{current_row}')
    total_cell = ws[f'E{current_row}']
    total_cell.value = f"총 {prev_total}명"
    total_cell.font = Font(bold=True)
    total_cell.alignment = center_align
    
    ws.merge_cells(f'U{current_row}:AF{current_row}')
    total_cell2 = ws[f'U{current_row}']
    total_cell2.value = f"총 {prev_total}명"
    total_cell2.font = Font(bold=True)
    total_cell2.alignment = center_align
    
    current_row += 2

# 분리 규칙 위반 명단
current_row += 2
ws.merge_cells(f'A{current_row}:F{current_row}')
violation_header = ws[f'A{current_row}']
violation_header.value = "분리 규칙 위반 명단 (같은 색 글자 = 같은 위반 쌍)"
violation_header.font = Font(bold=True, size=14)
violation_header.fill = header_fill
current_row += 1

if violations:
    headers = ['반', '학생1 학번', '학생1 이름', '학생2 학번', '학생2 이름', '불가피 여부']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    current_row += 1
    
    for v in violations:
        id1 = v['학생1']['학번']
        group_id = student_to_group.get(id1, -1)
        
        if group_id >= 0:
            color = violation_colors[group_id % len(violation_colors)]
        else:
            color = 'FF0000'
        
        ws.cell(row=current_row, column=1).value = v['반']
        ws.cell(row=current_row, column=2).value = v['학생1']['학번']
        ws.cell(row=current_row, column=3).value = v['학생1']['이름']
        ws.cell(row=current_row, column=4).value = v['학생2']['학번']
        ws.cell(row=current_row, column=5).value = v['학생2']['이름']
        ws.cell(row=current_row, column=6).value = "위반"
        
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_style
            cell.alignment = center_align
            if col <= 5:
                cell.font = Font(color=color, bold=True)
            else:
                cell.font = Font(bold=True) if v.get('inevitable', False) else Font()
        
        current_row += 1
else:
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "분리 규칙 위반 없음"
    cell.alignment = center_align
    cell.font = Font(bold=True)
    current_row += 1

# 수동 배정 필요한 학생 명단
current_row += 2
ws.merge_cells(f'A{current_row}:D{current_row}')
manual_header = ws[f'A{current_row}']
manual_header.value = "수동 배정 필요한 학생 명단 (separation.xlsx에 없는 학생)"
manual_header.font = Font(bold=True, size=14)
manual_header.fill = header_fill
current_row += 1

if len(df_manual_students) > 0:
    headers = ['학번', '이름', '이전학반', '남녀']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    current_row += 1
    
    for _, student in df_manual_students.iterrows():
        ws.cell(row=current_row, column=1).value = student['학번']
        ws.cell(row=current_row, column=2).value = student['이름']
        ws.cell(row=current_row, column=3).value = student['이전학반']
        ws.cell(row=current_row, column=4).value = student['남녀']
        
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_style
            cell.alignment = center_align
        
        current_row += 1
else:
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "수동 배정 필요한 학생 없음"
    cell.alignment = center_align
    cell.font = Font(bold=True)
    current_row += 1

# 최종 반별 요약
current_row += 2
ws.merge_cells(f'A{current_row}:D{current_row}')
summary_header = ws[f'A{current_row}']
summary_header.value = "최종 반별 구성 요약"
summary_header.font = Font(bold=True, size=14)
current_row += 2

summary_headers = ['반', '남학생', '여학생', '총학생수']
for col_idx, header in enumerate(summary_headers, 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border_style
current_row += 1

# transfer_plan 기반으로 각 반별 합계 계산 (각 이전학반별 테이블의 D열, T열 합계)
total_planned_male = 0
total_planned_female = 0
total_planned = 0

for target in target_classes:
    # transfer_plan에서 각 반별로 모든 이전학반에서 오는 학생 수 합계
    # (각 이전학반별 테이블의 D열, T열 값들의 합계)
    planned_male = sum(transfer_plan[prev_class]['male'][target] for prev_class in previous_classes)
    planned_female = sum(transfer_plan[prev_class]['female'][target] for prev_class in previous_classes)
    planned_total = planned_male + planned_female
    
    total_planned_male += planned_male
    total_planned_female += planned_female
    total_planned += planned_total
    
    # 목표 인원수와 비교
    target_male = target_distribution['male'][target]
    target_female = target_distribution['female'][target]
    target_total = target_distribution['total'][target]
    
    # 불일치 시 경고 출력
    if planned_male != target_male or planned_female != target_female:
        print(f"[경고] {target}반: transfer_plan 합계 {planned_total}명 (남 {planned_male}, 여 {planned_female}) != 목표 {target_total}명 (남 {target_male}, 여 {target_female})")
    
    ws.cell(row=current_row, column=1).value = f"{target}반"
    ws.cell(row=current_row, column=1).fill = colors[target]
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).value = planned_male
    ws.cell(row=current_row, column=3).value = planned_female
    ws.cell(row=current_row, column=4).value = planned_total
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = center_align
        cell.border = border_style
    
    current_row += 1

# 합계 행 추가
summary_headers_with_total = ['합계', '남학생', '여학생', '총학생수']
for col_idx, header in enumerate(summary_headers_with_total, 1):
    cell = ws.cell(row=current_row, column=col_idx)
    if col_idx == 1:
        cell.value = "합계"
    elif col_idx == 2:
        cell.value = total_planned_male
    elif col_idx == 3:
        cell.value = total_planned_female
    elif col_idx == 4:
        cell.value = total_planned
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = center_align
    cell.border = border_style

# 검증 출력
print(f"\n[최종 반별 구성 요약 검증]")
print(f"transfer_plan 합계: 총 {total_planned}명 (남 {total_planned_male}명, 여 {total_planned_female}명)")
print(f"목표 합계: 총 {total_students}명 (남 {male_count}명, 여 {female_count}명)")
if total_planned == total_students and total_planned_male == male_count and total_planned_female == female_count:
    print("[OK] 모든 숫자가 일치합니다!")
else:
    print("[X] 숫자가 일치하지 않습니다!")
    print(f"  차이: 총 {total_planned - total_students}명, 남 {total_planned_male - male_count}명, 여 {total_planned_female - female_count}명")

# 열 너비 조정
for col in ['A', 'B', 'Q', 'R']:
    ws.column_dimensions[col].width = 8
for col in ['C', 'S']:
    ws.column_dimensions[col].width = 10
for col in ['D', 'T']:
    ws.column_dimensions[col].width = 10
for col in range(5, 17):
    ws.column_dimensions[get_column_letter(col)].width = 12
for col in range(21, 33):
    ws.column_dimensions[get_column_letter(col)].width = 12

# 파일 저장
output_file = '반편성_배정표.xlsx'
try:
    wb.save(output_file)
    print(f"\n엑셀 파일이 생성되었습니다: {output_file}")
except PermissionError:
    import time
    output_file = f'반편성_배정표_{int(time.time())}.xlsx'
    wb.save(output_file)
    print(f"\n파일이 열려있어 다른 이름으로 저장했습니다: {output_file}")

if len(violations) > 0:
    print(f"\n위반 분석:")
    for v in violations:
        print(f"  {v['학생1']['이름']} - {v['학생2']['이름']} ({v['반']}반): 위반")

# 미배정 학생이 있으면 오류 코드로 종료
if unassigned_ids:
    print("\n" + "!" * 80)
    print(f"프로그램 종료: {len(unassigned_ids)}명의 학생이 배정되지 않았습니다.")
    print("!" * 80)
    import sys
    sys.exit(1)
