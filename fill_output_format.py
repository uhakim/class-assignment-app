# -*- coding: utf-8 -*-
"""
출력서식.xlsx에 반편성_완료.xlsx 데이터 채우기
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import re
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

# 올해 연도
current_year = 2026  # 2026학년도
current_grade = 4  # 기본값 (학번에서 추출)
previous_grade = 3  # 기본값 (학번에서 추출)
previous_year = 2025  # 2025학년도 (이전 학년도)

# 학년도 계산: 1~2월이면 현재 연도, 3~12월이면 현재 연도 + 1
current_month = datetime.now().month
if current_month >= 3:
    school_year = current_year + 1
else:
    school_year = current_year

print("=" * 80)
print("출력서식 채우기")
print("=" * 80)

# 반편성_완료.xlsx 읽기
print("\n반편성_완료.xlsx 읽는 중...")
df_completed = pd.read_excel('반편성_완료.xlsx', sheet_name='반편성 배정표')

# 학생자료.xlsx 읽기 (추가 정보용)
print("학생자료.xlsx 읽는 중...")
df_students = pd.read_excel('학생자료.xlsx')
student_info_dict = {}

# 학번 첫자리에서 현재 학년 추출
grade_from_student_id = None
for _, row in df_students.iterrows():
    student_id = str(row['학번'])
    if student_id and len(student_id) > 0 and student_id[0].isdigit():
        grade_from_student_id = int(student_id[0])
        break

if grade_from_student_id:
    current_grade = grade_from_student_id
    previous_grade = current_grade - 1
    print(f"학번 첫자리에서 학년 추출: {current_grade}학년 (이전 학년: {previous_grade}학년)")

for _, row in df_students.iterrows():
    student_id = str(row['학번'])
    student_info_dict[student_id] = {
        '이름': row['이름'],
        '생년월일': row.get('생년월일', ''),
        '학력수준': row.get('학력수준', ''),
        '영어반': row.get('영어반', ''),
        '제2외국어': row.get('제2외국어', ''),
        '비고': row.get('비고', ''),
        '남녀': row.get('남녀', '')
    }

# 데이터 파싱
print("\n데이터 파싱 중...")

# 이전 학반별 배정 데이터
previous_class_data = {
    1: {'A': {'male': [], 'female': []}, 'B': {'male': [], 'female': []}, 
        'C': {'male': [], 'female': []}, 'D': {'male': [], 'female': []}},
    2: {'A': {'male': [], 'female': []}, 'B': {'male': [], 'female': []}, 
        'C': {'male': [], 'female': []}, 'D': {'male': [], 'female': []}},
    3: {'A': {'male': [], 'female': []}, 'B': {'male': [], 'female': []}, 
        'C': {'male': [], 'female': []}, 'D': {'male': [], 'female': []}},
    4: {'A': {'male': [], 'female': []}, 'B': {'male': [], 'female': []}, 
        'C': {'male': [], 'female': []}, 'D': {'male': [], 'female': []}}
}

# 배정반별 데이터
target_class_data = {
    'A': {'male': [], 'female': []},
    'B': {'male': [], 'female': []},
    'C': {'male': [], 'female': []},
    'D': {'male': [], 'female': []}
}

# 데이터 파싱 - 1반도 포함하도록 수정
current_prev_class = None
for idx, row in df_completed.iterrows():
    # 이전 학반 헤더 확인 - 더 유연하게
    if pd.notna(row.iloc[0]):
        row0_str = str(row.iloc[0])
        if '학년' in row0_str and '반' in row0_str:
            if '1반' in row0_str or '1 반' in row0_str:
                current_prev_class = 1
            elif '2반' in row0_str or '2 반' in row0_str:
                current_prev_class = 2
            elif '3반' in row0_str or '3 반' in row0_str:
                current_prev_class = 3
            elif '4반' in row0_str or '4 반' in row0_str:
                current_prev_class = 4
            continue
    
    # 헤더 행 건너뛰기
    if pd.notna(row.iloc[0]) and str(row.iloc[0]) == '학년':
        continue
    
    # 합계 행 건너뛰기
    if pd.notna(row.iloc[2]) and '합계' in str(row.iloc[2]):
        continue
    
    # 데이터 행 처리 - 현재 학년이고 배정반이 A,B,C,D인 경우
    grade_val = row.iloc[0]
    prev_class_val = row.iloc[1] if pd.notna(row.iloc[1]) else None
    target_class_val = row.iloc[2] if pd.notna(row.iloc[2]) else None
    
    # 현재 학년이고, 배정반이 A, B, C, D 중 하나인 경우
    if pd.notna(grade_val) and (grade_val == current_grade or str(grade_val) == str(current_grade)):
        if pd.notna(target_class_val) and str(target_class_val) in ['A', 'B', 'C', 'D']:
            # 이전학반 확인
            if pd.notna(prev_class_val):
                try:
                    prev_class = int(prev_class_val)
                    if prev_class not in [1, 2, 3, 4]:
                        prev_class = current_prev_class if current_prev_class else 1
                except (ValueError, TypeError):
                    prev_class = current_prev_class if current_prev_class else 1
            else:
                prev_class = current_prev_class if current_prev_class else 1
            
            target_class = str(target_class_val)
            
            if target_class and target_class in ['A', 'B', 'C', 'D']:
                # 남학생 데이터
                for i in range(1, 7):
                    col_idx = 4 + (i-1) * 2
                    student_id = str(row.iloc[col_idx]) if pd.notna(row.iloc[col_idx]) else None
                    student_name = str(row.iloc[col_idx + 1]) if pd.notna(row.iloc[col_idx + 1]) else None
                    
                    if student_id and student_name and student_id != 'nan' and student_name != 'nan':
                        student_info = {
                            '학번': student_id,
                            '이름': student_name,
                            '이전학반': prev_class,
                            '배정반': target_class
                        }
                        # 학생자료에서 추가 정보 가져오기
                        if student_id in student_info_dict:
                            student_info.update(student_info_dict[student_id])
                        previous_class_data[prev_class][target_class]['male'].append(student_info)
                        target_class_data[target_class]['male'].append(student_info)
                
                # 여학생 데이터 (16번째 컬럼부터)
                for i in range(1, 7):
                    col_idx = 20 + (i-1) * 2
                    student_id = str(row.iloc[col_idx]) if pd.notna(row.iloc[col_idx]) else None
                    student_name = str(row.iloc[col_idx + 1]) if pd.notna(row.iloc[col_idx + 1]) else None
                    
                    if student_id and student_name and student_id != 'nan' and student_name != 'nan':
                        student_info = {
                            '학번': student_id,
                            '이름': student_name,
                            '이전학반': prev_class,
                            '배정반': target_class
                        }
                        # 학생자료에서 추가 정보 가져오기
                        if student_id in student_info_dict:
                            student_info.update(student_info_dict[student_id])
                        previous_class_data[prev_class][target_class]['female'].append(student_info)
                        target_class_data[target_class]['female'].append(student_info)
        else:
            continue
    else:
        continue

# 학번순 정렬
for prev_class in [1, 2, 3, 4]:
    for target_class in ['A', 'B', 'C', 'D']:
        previous_class_data[prev_class][target_class]['male'].sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)
        previous_class_data[prev_class][target_class]['female'].sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)

for target_class in ['A', 'B', 'C', 'D']:
    target_class_data[target_class]['male'].sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)
    target_class_data[target_class]['female'].sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)

print(f"\n데이터 파싱 완료:")
for prev_class in [1, 2, 3, 4]:
    for target_class in ['A', 'B', 'C', 'D']:
        male_count = len(previous_class_data[prev_class][target_class]['male'])
        female_count = len(previous_class_data[prev_class][target_class]['female'])
        if male_count > 0 or female_count > 0:
            print(f"  {prev_class}반 -> {target_class}반: 남 {male_count}명, 여 {female_count}명")

# 출력서식.xlsx 열기
print("\n출력서식.xlsx 열기...")
wb = load_workbook('출력서식.xlsx')

# 년도 업데이트 함수
def update_year_in_cell(cell_value, year, grade):
    """셀 값에서 년도와 학년 업데이트"""
    if cell_value is None:
        return None
    cell_str = str(cell_value)
    # 2025학년도 -> 현재년도학년도
    cell_str = re.sub(r'\d{4}학년도', f'{year}학년도', cell_str)
    # (     )학년 -> ( 4 )학년
    cell_str = re.sub(r'\([\s]*\)\s*학년', f'( {grade} )학년', cell_str)
    return cell_str

# 영어반 분류 함수
def classify_english_level(english_class):
    """영어반을 E1/E2, E3/E4, E5/E6, E7/E8로 분류"""
    if pd.isna(english_class) or english_class == '':
        return None, None, None, None
    eng_str = str(english_class).upper()
    e1_e2 = 1 if 'E1' in eng_str or 'E2' in eng_str else 0
    e3_e4 = 1 if 'E3' in eng_str or 'E4' in eng_str else 0
    e5_e6 = 1 if 'E5' in eng_str or 'E6' in eng_str else 0
    e7_e8 = 1 if 'E7' in eng_str or 'E8' in eng_str else 0
    return e1_e2, e3_e4, e5_e6, e7_e8

# 이전 학반별 시트 채우기
print("\n이전 학반별 시트 채우는 중...")
for prev_class in [1, 2, 3, 4]:
    for gender in ['남', '여']:
        sheet_name = f'{prev_class}반 {gender}'
        if sheet_name not in wb.sheetnames:
            print(f"  경고: {sheet_name} 시트를 찾을 수 없습니다.")
            continue
        
        ws = wb[sheet_name]
        print(f"  {sheet_name} 시트 처리 중...")
        
        # 년도 업데이트 (16행(P행)은 제외)
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            # 16행(P행)은 건드리지 않음
            if row_idx == 16:
                continue
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if '학년도' in cell.value or '학년' in cell.value:
                        new_value = update_year_in_cell(cell.value, current_year, current_grade)
                        if new_value != cell.value:
                            cell.value = new_value
        
        # 각 배정반(A, B, C, D) 섹션 찾아서 채우기
        target_classes = ['A', 'B', 'C', 'D']
        section_start_rows = {}
        
        # 섹션 시작 행 찾기 - 더 정확한 방법
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value and isinstance(cell_value, str):
                # "학생 A", "학생 B" 등의 패턴 찾기
                for target_class in target_classes:
                    if f'학생 {target_class}' in cell_value:
                        # 다음 몇 행에서 "순" 헤더 찾기
                        for search_row in range(row_idx + 1, min(row_idx + 5, ws.max_row + 1)):
                            header_cell = ws.cell(row=search_row, column=2).value
                            if header_cell and '순' in str(header_cell):
                                # 헤더 다음 행이 데이터 시작
                                section_start_rows[target_class] = search_row + 2
                                break
                        break
        
        # 데이터 채우기
        for target_class in target_classes:
            if target_class not in section_start_rows:
                # 섹션을 찾지 못한 경우, 패턴으로 다시 찾기
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=2).value
                    if cell_value and isinstance(cell_value, str):
                        if f'학생 {target_class}' in cell_value or (f'{target_class}' in cell_value and '학생' in cell_value):
                            # 다음 빈 행부터 데이터 시작
                            for data_row in range(row_idx + 3, min(row_idx + 20, ws.max_row + 1)):
                                if ws.cell(row=data_row, column=2).value is None or ws.cell(row=data_row, column=2).value == '':
                                    section_start_rows[target_class] = data_row
                                    break
                            break
            
            if target_class in section_start_rows:
                start_row = section_start_rows[target_class]
                # gender를 'male'/'female'로 변환
                gender_key = 'male' if gender == '남' else 'female'
                students = previous_class_data[prev_class][target_class][gender_key]
                
                # 학생 데이터 채우기
                for idx, student in enumerate(students):
                    row = start_row + idx
                    if row > ws.max_row:
                        break
                    
                    # 순번
                    ws.cell(row=row, column=2).value = idx + 1
                    # 이름
                    ws.cell(row=row, column=3).value = student.get('이름', '')
                    # 생년월일
                    birth = student.get('생년월일', '')
                    if pd.notna(birth) and birth != '':
                        ws.cell(row=row, column=4).value = str(birth)
                    # 성적 (학력수준)
                    ws.cell(row=row, column=5).value = student.get('학력수준', '')
                    # 영어
                    ws.cell(row=row, column=6).value = student.get('영어반', '')
                    # 중/일
                    ws.cell(row=row, column=7).value = student.get('제2외국어', '')
                    # 비고
                    ws.cell(row=row, column=8).value = student.get('비고', '')
        
        # n반 남/여 시트의 2행, 16행, 30행, 44행에 (     )학년도 (     )학년 (     )반  남학생 A,B,C,D 형식으로 반 정보 채우기
        # 2행: A반, 16행: B반, 30행: C반, 44행: D반
        # 16열(P열)은 완전히 비워야 함
        print(f"    반 정보 채우기: 시트={sheet_name}, prev_class={prev_class}, gender={gender}")
        gender_text = '남학생' if gender == '남' else '여학생'
        
        # 병합된 셀 범위 확인
        merged_ranges = list(ws.merged_cells.ranges)
        
        def get_merged_cell_coord(row, col):
            """병합된 셀의 경우 첫 번째 셀의 좌표 반환"""
            for merged_range in merged_ranges:
                if row >= merged_range.min_row and row <= merged_range.max_row and \
                   col >= merged_range.min_col and col <= merged_range.max_col:
                    return (merged_range.min_row, merged_range.min_col)
            return (row, col)
        
        def safe_write_cell(row, col, value):
            """병합된 셀을 고려하여 안전하게 셀에 값 쓰기"""
            try:
                cell = ws.cell(row=row, column=col)
                # MergedCell인지 확인
                if hasattr(cell, 'value') and hasattr(cell, 'coordinate'):
                    # 병합된 셀의 첫 번째 셀에만 쓰기
                    actual_row, actual_col = get_merged_cell_coord(row, col)
                    ws.cell(row=actual_row, column=actual_col).value = value
                else:
                    ws.cell(row=row, column=col).value = value
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀 찾아서 쓰기
                actual_row, actual_col = get_merged_cell_coord(row, col)
                ws.cell(row=actual_row, column=actual_col).value = value
        
        # 2행: A반 (모든 열에 채우기, 단 A열, I열, J열, K열, P열 제외)
        row_2 = 2
        class_A = 'A'
        text_row_2 = f'( {school_year} )학년도 ( {current_grade} )학년 ( {prev_class} )반  {gender_text} {class_A}'
        excluded_cols = [1, 9, 10, 11, 16]  # A열, I열, J열, K열, P열 제외
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in excluded_cols:
                continue
            # 기존 값이 반 정보 관련이면 교체
            try:
                val = ws.cell(row=row_2, column=col_idx).value
                if val is not None:
                    s = str(val)
                    if '학생' in s or '학년도' in s or ('학년' in s and '반' in s):
                        safe_write_cell(row_2, col_idx, text_row_2)
                else:
                    # 빈 셀이면 채우기
                    safe_write_cell(row_2, col_idx, text_row_2)
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀에만 쓰기
                actual_row, actual_col = get_merged_cell_coord(row_2, col_idx)
                if actual_row == row_2 and actual_col == col_idx:  # 첫 번째 셀인 경우만
                    ws.cell(row=actual_row, column=actual_col).value = text_row_2
        print(f"      2행 설정 완료: '{text_row_2}'")
        
        # 16행: B반 (모든 열에 채우기, 단 A열, I열, J열, K열, P열 제외)
        row_16 = 16
        class_B = 'B'
        text_row_16 = f'( {school_year} )학년도 ( {current_grade} )학년 ( {prev_class} )반  {gender_text} {class_B}'
        excluded_cols = [1, 9, 10, 11, 16]  # A열, I열, J열, K열, P열 제외
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in excluded_cols:
                continue
            # 기존 값이 반 정보 관련이면 교체
            try:
                val = ws.cell(row=row_16, column=col_idx).value
                if val is not None:
                    s = str(val)
                    if '학생' in s or '학년도' in s or ('학년' in s and '반' in s):
                        safe_write_cell(row_16, col_idx, text_row_16)
                else:
                    # 빈 셀이면 채우기
                    safe_write_cell(row_16, col_idx, text_row_16)
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀에만 쓰기
                actual_row, actual_col = get_merged_cell_coord(row_16, col_idx)
                if actual_row == row_16 and actual_col == col_idx:  # 첫 번째 셀인 경우만
                    ws.cell(row=actual_row, column=actual_col).value = text_row_16
        print(f"      16행 설정 완료: '{text_row_16}'")
        
        # 30행: C반 (모든 열에 채우기, 단 A열, I열, J열, K열, P열 제외)
        row_30 = 30
        class_C = 'C'
        text_row_30 = f'( {school_year} )학년도 ( {current_grade} )학년 ( {prev_class} )반  {gender_text} {class_C}'
        excluded_cols = [1, 9, 10, 11, 16]  # A열, I열, J열, K열, P열 제외
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in excluded_cols:
                continue
            # 기존 값이 반 정보 관련이면 교체
            try:
                val = ws.cell(row=row_30, column=col_idx).value
                if val is not None:
                    s = str(val)
                    if '학생' in s or '학년도' in s or ('학년' in s and '반' in s):
                        safe_write_cell(row_30, col_idx, text_row_30)
                else:
                    # 빈 셀이면 채우기
                    safe_write_cell(row_30, col_idx, text_row_30)
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀에만 쓰기
                actual_row, actual_col = get_merged_cell_coord(row_30, col_idx)
                if actual_row == row_30 and actual_col == col_idx:  # 첫 번째 셀인 경우만
                    ws.cell(row=actual_row, column=actual_col).value = text_row_30
        print(f"      30행 설정 완료: '{text_row_30}'")
        
        # 44행: D반 (모든 열에 채우기, 단 A열, I열, J열, K열, P열 제외)
        row_44 = 44
        class_D = 'D'
        text_row_44 = f'( {school_year} )학년도 ( {current_grade} )학년 ( {prev_class} )반  {gender_text} {class_D}'
        excluded_cols = [1, 9, 10, 11, 16]  # A열, I열, J열, K열, P열 제외
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in excluded_cols:
                continue
            # 기존 값이 반 정보 관련이면 교체
            try:
                val = ws.cell(row=row_44, column=col_idx).value
                if val is not None:
                    s = str(val)
                    if '학생' in s or '학년도' in s or ('학년' in s and '반' in s):
                        safe_write_cell(row_44, col_idx, text_row_44)
                else:
                    # 빈 셀이면 채우기
                    safe_write_cell(row_44, col_idx, text_row_44)
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀에만 쓰기
                actual_row, actual_col = get_merged_cell_coord(row_44, col_idx)
                if actual_row == row_44 and actual_col == col_idx:  # 첫 번째 셀인 경우만
                    ws.cell(row=actual_row, column=actual_col).value = text_row_44
        print(f"      44행 설정 완료: '{text_row_44}'")
        
        # 16열(P열)의 모든 행에서 반 정보 제거 (완전히 비우기)
        for row_idx in range(1, ws.max_row + 1):
            try:
                val = ws.cell(row=row_idx, column=16).value
                if val is not None:
                    s = str(val)
                    if '학년도' in s or ('학년' in s and '반' in s) or '학생' in s:
                        actual_row, actual_col = get_merged_cell_coord(row_idx, 16)
                        if actual_row == row_idx and actual_col == 16:  # 첫 번째 셀인 경우만
                            ws.cell(row=actual_row, column=actual_col).value = None
                        print(f"        16열 {row_idx}행: 반 정보 제거")
            except AttributeError:
                # MergedCell인 경우 첫 번째 셀에만 쓰기
                actual_row, actual_col = get_merged_cell_coord(row_idx, 16)
                if actual_row == row_idx and actual_col == 16:  # 첫 번째 셀인 경우만
                    ws.cell(row=actual_row, column=actual_col).value = None

# 학급별 시트 채우기
print("\n학급별 시트 채우는 중...")
for prev_class_num in [1, 2, 3, 4]:
    sheet_name = f'학급별 {prev_class_num}반'
    
    if sheet_name not in wb.sheetnames:
        print(f"  경고: {sheet_name} 시트를 찾을 수 없습니다.")
        continue
    
    ws = wb[sheet_name]
    print(f"  {sheet_name} 시트 처리 중...")
    
    # 년도 업데이트
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '학년도' in cell.value or '학년' in cell.value:
                    new_value = update_year_in_cell(cell.value, current_year, current_grade)
                    if new_value != cell.value:
                        cell.value = new_value
                    # 반 정보 채우기
                    if '(     )' in new_value:
                        new_value = new_value.replace('(     )', f'({prev_class_num})')
                        cell.value = new_value
    
    # 진급 이전의 해당 학반 학생 데이터 채우기 (1반, 2반, 3반, 4반 각각)
    # 이전학반이 prev_class_num인 모든 학생들 (A, B, C, D 어디로든 배정된 학생들)
    all_students = []
    for target_class in ['A', 'B', 'C', 'D']:
        all_students.extend(previous_class_data[prev_class_num][target_class]['male'])
        all_students.extend(previous_class_data[prev_class_num][target_class]['female'])
    
    # 학번순 정렬
    all_students.sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)
    
    # 데이터 시작 행 찾기 (번호 컬럼이 있는 행)
    start_row = 5  # 일반적으로 5행부터 시작
    for row_idx in range(1, 10):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value and '번호' in str(cell_value):
            start_row = row_idx + 2
            break
    
    # 학생 데이터 채우기 - B열부터 번호 시작 (한 칸 우측으로 이동)
    # 1~20번: 기존 위치 (B~J열)
    # 41~60번: 25열(Y열)부터 시작
    max_first_section = 20  # 첫 번째 섹션에 들어갈 최대 학생 수 (1~20번)
    second_section_start_number = 41  # 두 번째 섹션 시작 번호
    second_section_end_number = 60  # 두 번째 섹션 종료 번호
    
    for idx, student in enumerate(all_students):
        # 번호 - 학번의 뒤 두 자리
        student_id = str(student.get('학번', ''))
        if len(student_id) >= 2:
            number_str = student_id[-2:]  # 뒤 두 자리
            try:
                number = int(number_str)
            except:
                continue  # 번호가 없으면 입력 안 함
        else:
            continue  # 번호가 없으면 입력 안 함
        
        if number <= max_first_section:
            # 1~20번: 기존 위치
            row = start_row + (number - 1)
            if row > ws.max_row:
                break
            
            ws.cell(row=row, column=2).value = number
            ws.cell(row=row, column=3).value = student.get('이름', '')
            birth = student.get('생년월일', '')
            if pd.notna(birth) and birth != '':
                ws.cell(row=row, column=4).value = str(birth)
            ws.cell(row=row, column=5).value = student.get('학력수준', '')
            ws.cell(row=row, column=6).value = student.get('영어반', '')
            ws.cell(row=row, column=7).value = student.get('제2외국어', '')
            ws.cell(row=row, column=8).value = student.get('비고', '')
            ws.cell(row=row, column=9).value = student.get('배정반', '')
        elif second_section_start_number <= number <= second_section_end_number:
            # 41~60번: B25(2열, 25행)부터 시작
            second_section_start_row = 25  # 25행부터 시작
            row = second_section_start_row + (number - second_section_start_number)
            if row > ws.max_row:
                break
            
            ws.cell(row=row, column=2).value = number  # 번호 (B열)
            ws.cell(row=row, column=3).value = student.get('이름', '')  # 이름
            birth = student.get('생년월일', '')
            if pd.notna(birth) and birth != '':
                ws.cell(row=row, column=4).value = str(birth)  # 생년월일
            ws.cell(row=row, column=5).value = student.get('학력수준', '')  # 성적
            ws.cell(row=row, column=6).value = student.get('영어반', '')  # 영어
            ws.cell(row=row, column=7).value = student.get('제2외국어', '')  # 중/일/스
            ws.cell(row=row, column=8).value = student.get('비고', '')  # 특이사항
            ws.cell(row=row, column=9).value = student.get('배정반', '')  # 그룹
    
    # 학급별 n반 시트 2열에 ( 4 )학년 ( 1 ) 반 형식으로 띄어쓰기 수정
    for row_idx in range(1, 5):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value and isinstance(cell_value, str):
            cell_str = str(cell_value)
            # ( 4 )학년 (1) 반 형식을 ( 4 )학년 ( 1 ) 반 형식으로 수정
            if '( 4 )' in cell_str or '(4)' in cell_str:
                new_value = cell_str.replace('( 4 )', f'( {current_grade} )').replace('(4)', f'( {current_grade} )')
                ws.cell(row=row_idx, column=2).value = new_value
            if '(     )' in cell_str:
                new_value = cell_str.replace('(     )', f'( {prev_class_num} )')
                ws.cell(row=row_idx, column=2).value = new_value
            elif '(  )' in cell_str:
                new_value = cell_str.replace('(  )', f'( {prev_class_num} )')
                ws.cell(row=row_idx, column=2).value = new_value
            elif '(1)' in cell_str or '(2)' in cell_str or '(3)' in cell_str or '(4)' in cell_str:
                # (1) 형식을 ( 1 ) 형식으로 변경
                new_value = re.sub(r'\((\d+)\)', r'( \1 )', cell_str)
                ws.cell(row=row_idx, column=2).value = new_value

# 통계표 시트 채우기
print("\n통계표 시트 채우는 중...")
if '통계표' in wb.sheetnames:
    ws = wb['통계표']
    
    # 년도 업데이트 및 B열의 ○ 교체
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '2025년' in cell.value:
                    cell.value = cell.value.replace('2025년', f'{current_year}년')
                if '2024년' in cell.value:
                    cell.value = cell.value.replace('2024년', f'{current_year - 1}년')
                # B열(2열)에서 ○의 N 형식을 current_grade의 N으로 교체
                if cell.column == 2 and '○의' in cell.value:
                    # ○의 1, ○의 2 등을 current_grade의 1, current_grade의 2로 교체
                    new_value = cell.value.replace('○의', f'{current_grade}의')
                    cell.value = new_value
    
    # 신 학반과 현 학반 정보 채우기
    # 병합된 셀 범위 확인
    merged_ranges = list(ws.merged_cells.ranges)
    
    def get_merged_cell_value(row, col):
        """병합된 셀의 경우 첫 번째 셀의 좌표 반환"""
        for merged_range in merged_ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                return (merged_range.min_row, merged_range.min_col)
        return (row, col)
    
    def safe_write_stat_cell(row, col, value):
        """통계표에서 병합된 셀을 고려하여 안전하게 쓰기"""
        merge_row, merge_col = get_merged_cell_value(row, col)
        if merge_row == row and merge_col == col:  # 첫 번째 셀인 경우만 쓰기
            ws.cell(row=row, column=col).value = value
    
    # 통계표 구조:
    # 행 3~6: A반의 1~4반 데이터 (C3부터 데이터 시작)
    # 행 7: A반 소계
    # 행 8~11: B반의 1~4반 데이터
    # 행 12: B반 소계
    # 행 13~16: C반의 1~4반 데이터
    # 행 17: C반 소계
    # 행 18~21: D반의 1~4반 데이터
    # 행 22: D반 소계
    # 행 23: 총계
    
    # 각 배정반별 시작 행 계산 (행 3부터 시작)
    target_class_rows = {
        'A': {'start': 3, 'summary': 7},   # A반: 3~6행 데이터, 7행 소계
        'B': {'start': 8, 'summary': 12},   # B반: 8~11행 데이터, 12행 소계
        'C': {'start': 13, 'summary': 17},  # C반: 13~16행 데이터, 17행 소계
        'D': {'start': 18, 'summary': 22}   # D반: 18~21행 데이터, 22행 소계
    }
    
    # B열의 특정 행들(B3~6, B8~11, B13~16, B18~21)에서 ○의 N을 현재 학년으로 교체
    # 병합된 셀을 고려하여 처리
    b_rows_to_update = [3, 4, 5, 6, 8, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21]
    processed_merge_cells = set()  # 이미 처리한 병합 셀 추적
    
    for row_num in b_rows_to_update:
        if row_num <= ws.max_row:
            try:
                # 병합된 셀 확인
                merge_row, merge_col = get_merged_cell_value(row_num, 2)
                
                # 병합된 셀의 첫 번째 셀만 처리 (중복 처리 방지)
                if (merge_row, merge_col) not in processed_merge_cells:
                    processed_merge_cells.add((merge_row, merge_col))
                    
                    # 첫 번째 셀에 직접 접근
                    first_cell = ws.cell(row=merge_row, column=merge_col)
                    cell_value = first_cell.value
                    
                    if cell_value:
                        if isinstance(cell_value, str):
                            if '○의' in cell_value or '○' in cell_value:
                                # ○의 1, ○의 2 등을 current_grade의 1, current_grade의 2로 교체
                                new_value = cell_value.replace('○의', f'{current_grade}의').replace('○', str(current_grade))
                                first_cell.value = new_value
                                print(f"  B{merge_row} (병합 셀): '{cell_value}' -> '{new_value}'")
                        # 숫자나 다른 타입은 건너뜀
            except Exception as e:
                print(f"  B{row_num} 처리 중 오류: {e}")
                pass
    
    # 추가로 B열 전체를 스캔하여 ○가 있는 모든 셀 처리 (안전장치)
    for row_idx in range(1, ws.max_row + 1):
        try:
            cell = ws.cell(row=row_idx, column=2)
            cell_value = cell.value
            if cell_value and isinstance(cell_value, str):
                if '○의' in cell_value:
                    new_value = cell_value.replace('○의', f'{current_grade}의')
                    cell.value = new_value
                    print(f"  B{row_idx} (전체 스캔): '{cell_value}' -> '{new_value}'")
        except:
            pass
    
    for target_class in ['A', 'B', 'C', 'D']:
        data_start_row = target_class_rows[target_class]['start']
        summary_row = target_class_rows[target_class]['summary']
        
        # A열: A3에 "○의 A"가 있으면 "4의 A"로 변경 (각 배정반의 첫 번째 데이터 행만)
        # A열은 대부분 건드리지 않고, A3, A8, A13, A18만 수정
        first_data_row = data_start_row
        if first_data_row <= ws.max_row:
            cell_value = ws.cell(row=first_data_row, column=1).value
            if cell_value and isinstance(cell_value, str):
                if '○의' in cell_value:
                    if target_class in cell_value:
                        try:
                            merge_row, merge_col = get_merged_cell_value(first_data_row, 1)
                            if merge_row == first_data_row and merge_col == 1:
                                ws.cell(row=first_data_row, column=1).value = f'{current_grade + 1}의 {target_class}'
                        except:
                            pass
        
        for prev_class_idx, prev_class in enumerate([1, 2, 3, 4]):
            row = data_start_row + prev_class_idx
            
            # B열: B3부터 "3의 1", "3의 2" 형식으로 입력
            # B1, B2는 건드리지 않음, B3부터만 수정
            if row >= 3:  # B3부터만 수정
                try:
                    merge_row, merge_col = get_merged_cell_value(row, 2)
                    if merge_row == row and merge_col == 2:  # 첫 번째 셀인 경우만
                        # 기존 값이 "○의"로 시작하거나 비어있을 때만 수정
                        existing_value = ws.cell(row=row, column=2).value
                        if existing_value is None or (isinstance(existing_value, str) and ('○의' in existing_value or existing_value.strip() == '' or '2의' in existing_value)):
                            # ○의가 있으면 current_grade로 교체
                            if existing_value and '○의' in existing_value:
                                ws.cell(row=row, column=2).value = existing_value.replace('○의', f'{current_grade}의')
                            else:
                                ws.cell(row=row, column=2).value = f'{current_grade}의 {prev_class}'
                except:
                    pass
            
            # V열(22열): 현학반에 있는 ○를 학번 첫자리로 교체
            # 학생자료에서 해당 이전학반의 학생들의 학번 첫자리를 확인
            if row <= ws.max_row:
                try:
                    v_cell_value = ws.cell(row=row, column=22).value  # V열은 22번째 열
                    if v_cell_value and isinstance(v_cell_value, str) and '○' in v_cell_value:
                        # 해당 이전학반의 학생들의 학번 첫자리 확인
                        sample_students = previous_class_data[prev_class][target_class]['male'] + previous_class_data[prev_class][target_class]['female']
                        if len(sample_students) > 0:
                            # 첫 번째 학생의 학번 첫자리 사용
                            first_student_id = sample_students[0].get('학번', '')
                            if first_student_id and len(str(first_student_id)) > 0:
                                grade_from_id = str(first_student_id)[0]  # 학번 첫자리
                                # ○를 학번 첫자리로 교체
                                new_v_value = v_cell_value.replace('○', grade_from_id)
                                merge_row, merge_col = get_merged_cell_value(row, 22)
                                if merge_row == row and merge_col == 22:
                                    ws.cell(row=row, column=22).value = new_v_value
                except:
                    pass
            
            # 학생 데이터 가져오기
            students = previous_class_data[prev_class][target_class]['male'] + previous_class_data[prev_class][target_class]['female']
            
            if len(students) > 0:
                # 영어반 통계 (E1/E2, E3/E4, E5/E6, E7/E8)
                e1_e2_count = 0
                e3_e4_count = 0
                e5_e6_count = 0
                e7_e8_count = 0
                e1_count = 0
                e3_count = 0
                e5_count = 0
                e7_count = 0
                e2_count = 0
                e4_count = 0
                e6_count = 0
                e8_count = 0
                
                # 제2외국어 통계
                chinese_count = 0
                japanese_count = 0
                
                # 학력수준 통계
                level_4_count = 0
                level_3_count = 0
                level_2_count = 0
                level_1_count = 0
                level_p_count = 0
                
                # 남녀 통계
                male_count = len(previous_class_data[prev_class][target_class]['male'])
                female_count = len(previous_class_data[prev_class][target_class]['female'])
                
                for student in students:
                    # 영어반 분류
                    eng = student.get('영어반', '')
                    if pd.notna(eng) and eng != '':
                        eng_str = str(eng).upper()
                        if 'E1' in eng_str:
                            e1_count += 1
                            e1_e2_count += 1
                        elif 'E2' in eng_str:
                            e2_count += 1
                            e1_e2_count += 1
                        elif 'E3' in eng_str:
                            e3_count += 1
                            e3_e4_count += 1
                        elif 'E4' in eng_str:
                            e4_count += 1
                            e3_e4_count += 1
                        elif 'E5' in eng_str:
                            e5_count += 1
                            e5_e6_count += 1
                        elif 'E6' in eng_str:
                            e6_count += 1
                            e5_e6_count += 1
                        elif 'E7' in eng_str:
                            e7_count += 1
                            e7_e8_count += 1
                        elif 'E8' in eng_str:
                            e8_count += 1
                            e7_e8_count += 1
                    
                    # 제2외국어
                    lang = student.get('제2외국어', '')
                    if pd.notna(lang) and lang != '':
                        lang_str = str(lang)
                        if '중' in lang_str or '중국' in lang_str:
                            chinese_count += 1
                        elif '일' in lang_str or '일본' in lang_str:
                            japanese_count += 1
                    
                    # 학력수준
                    level = student.get('학력수준', '')
                    if pd.notna(level) and level != '':
                        level_str = str(level)
                        if '4' in level_str or '상' in level_str:
                            level_4_count += 1
                        elif '3' in level_str or '중상' in level_str:
                            level_3_count += 1
                        elif '2' in level_str or '중' in level_str:
                            level_2_count += 1
                        elif '1' in level_str or '하' in level_str:
                            level_1_count += 1
                        elif 'P' in level_str.upper() or '부진' in level_str:
                            level_p_count += 1
                
                # 영어반 통계 입력 (C열부터 시작)
                # C열: E1/E2, D열: E3/E4, E열: E5/E6, F열: E7/E8
                # G열: 합계
                safe_write_stat_cell(row, 3, e1_e2_count)  # C열
                safe_write_stat_cell(row, 4, e3_e4_count)  # D열
                safe_write_stat_cell(row, 5, e5_e6_count)  # E열
                safe_write_stat_cell(row, 6, e7_e8_count)  # F열
                safe_write_stat_cell(row, 7, e1_e2_count + e3_e4_count + e5_e6_count + e7_e8_count)  # G열 합계
                
                # 제2외국어 통계
                # H열: 중국어, I열: 일본어, J열: 합계
                safe_write_stat_cell(row, 8, chinese_count)  # H열
                safe_write_stat_cell(row, 9, japanese_count)  # I열
                safe_write_stat_cell(row, 10, chinese_count + japanese_count)  # J열 합계
                
                # 학력수준 통계 (K열부터)
                safe_write_stat_cell(row, 11, level_4_count)  # K열
                safe_write_stat_cell(row, 12, level_3_count)  # L열
                safe_write_stat_cell(row, 13, level_2_count)  # M열
                safe_write_stat_cell(row, 14, level_1_count)  # N열
                safe_write_stat_cell(row, 15, level_p_count)  # O열
                safe_write_stat_cell(row, 16, level_4_count + level_3_count + level_2_count + level_1_count + level_p_count)  # P열
                
                # 재적수 (남여)
                safe_write_stat_cell(row, 17, male_count)  # Q열
                safe_write_stat_cell(row, 18, female_count)  # R열
                safe_write_stat_cell(row, 19, male_count + female_count)  # S열
                
                # 비고는 비워두기 (아무것도 입력하지 않음)
        
        # 각 배정반별 소계 계산 (행 9, 14, 19, 24)
        if summary_row <= ws.max_row:
            # 해당 배정반의 1~4반 데이터 합산
            # A반: 3~6행, B반: 10~13행, C반: 15~18행, D반: 20~23행
            # data_start_row는 이미 위에서 계산됨
            
            # 영어반 합계 (E~H열의 합을 I열에, E~H열 각각의 합)
            e_sum = 0
            f_sum = 0
            g_sum = 0
            h_sum = 0
            # 제2외국어 합계 (J, K열의 합을 L열에)
            j_sum = 0
            k_sum = 0
            # 학력수준 합계 (M~Q열의 합을 R열에)
            m_sum = 0
            n_sum = 0
            o_sum = 0
            p_sum = 0
            q_sum = 0
            # 재적수 합계
            s_sum = 0
            t_sum = 0
            u_sum = 0
            
            # 해당 배정반의 모든 이전 학반 데이터 합산
            for prev_class in [1, 2, 3, 4]:
                students = previous_class_data[prev_class][target_class]['male'] + previous_class_data[prev_class][target_class]['female']
                for student in students:
                    eng = student.get('영어반', '')
                    if pd.notna(eng) and eng != '':
                        eng_str = str(eng).upper()
                        if 'E1' in eng_str or 'E2' in eng_str:
                            e_sum += 1
                        elif 'E3' in eng_str or 'E4' in eng_str:
                            f_sum += 1
                        elif 'E5' in eng_str or 'E6' in eng_str:
                            g_sum += 1
                        elif 'E7' in eng_str or 'E8' in eng_str:
                            h_sum += 1
                    lang = student.get('제2외국어', '')
                    if pd.notna(lang) and lang != '':
                        lang_str = str(lang)
                        if '중' in lang_str or '중국' in lang_str:
                            j_sum += 1
                        elif '일' in lang_str or '일본' in lang_str:
                            k_sum += 1
                    level = student.get('학력수준', '')
                    if pd.notna(level) and level != '':
                        level_str = str(level)
                        if '4' in level_str or '상' in level_str:
                            m_sum += 1
                        elif '3' in level_str or '중상' in level_str:
                            n_sum += 1
                        elif '2' in level_str or '중' in level_str:
                            o_sum += 1
                        elif '1' in level_str or '하' in level_str:
                            p_sum += 1
                        elif 'P' in level_str.upper() or '부진' in level_str:
                            q_sum += 1
                
                male_total = len(previous_class_data[prev_class][target_class]['male'])
                female_total = len(previous_class_data[prev_class][target_class]['female'])
                s_sum += male_total
                t_sum += female_total
                u_sum += male_total + female_total
            
            # 소계 입력 (C열부터 시작) - 병합된 셀 처리
            # C열: data_start_row부터 4행의 합을 summary_row에
            c_cell_sum = sum(ws.cell(row=r, column=3).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 3, c_cell_sum if c_cell_sum > 0 else e_sum)
            
            # D열: data_start_row부터 4행의 합을 summary_row에
            d_cell_sum = sum(ws.cell(row=r, column=4).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 4, d_cell_sum if d_cell_sum > 0 else f_sum)
            
            # E열: data_start_row부터 4행의 합을 summary_row에
            e_cell_sum = sum(ws.cell(row=r, column=5).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 5, e_cell_sum if e_cell_sum > 0 else g_sum)
            
            # F열: data_start_row부터 4행의 합을 summary_row에
            f_cell_sum = sum(ws.cell(row=r, column=6).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 6, f_cell_sum if f_cell_sum > 0 else h_sum)
            
            # G열: C~F열의 합계
            safe_write_stat_cell(summary_row, 7, c_cell_sum + d_cell_sum + e_cell_sum + f_cell_sum if (c_cell_sum + d_cell_sum + e_cell_sum + f_cell_sum) > 0 else (e_sum + f_sum + g_sum + h_sum))
            
            # 제2외국어 합계 (H~J열)
            # H열: data_start_row부터 4행의 합을 summary_row에
            h_cell_sum = sum(ws.cell(row=r, column=8).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 8, h_cell_sum if h_cell_sum > 0 else j_sum)
            
            # I열: data_start_row부터 4행의 합을 summary_row에
            i_cell_sum = sum(ws.cell(row=r, column=9).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            safe_write_stat_cell(summary_row, 9, i_cell_sum if i_cell_sum > 0 else k_sum)
            
            # J열: H~I열의 합계
            safe_write_stat_cell(summary_row, 10, h_cell_sum + i_cell_sum if (h_cell_sum + i_cell_sum) > 0 else (j_sum + k_sum))
            
            # 학력수준 합계 (K~P열)
            k_cell_sum = sum(ws.cell(row=r, column=11).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            l_cell_sum = sum(ws.cell(row=r, column=12).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            m_cell_sum = sum(ws.cell(row=r, column=13).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            n_cell_sum = sum(ws.cell(row=r, column=14).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            o_cell_sum = sum(ws.cell(row=r, column=15).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            
            safe_write_stat_cell(summary_row, 11, k_cell_sum if k_cell_sum > 0 else m_sum)
            safe_write_stat_cell(summary_row, 12, l_cell_sum if l_cell_sum > 0 else n_sum)
            safe_write_stat_cell(summary_row, 13, m_cell_sum if m_cell_sum > 0 else o_sum)
            safe_write_stat_cell(summary_row, 14, n_cell_sum if n_cell_sum > 0 else p_sum)
            safe_write_stat_cell(summary_row, 15, o_cell_sum if o_cell_sum > 0 else q_sum)
            safe_write_stat_cell(summary_row, 16, k_cell_sum + l_cell_sum + m_cell_sum + n_cell_sum + o_cell_sum if (k_cell_sum + l_cell_sum + m_cell_sum + n_cell_sum + o_cell_sum) > 0 else (m_sum + n_sum + o_sum + p_sum + q_sum))
            
            # 재적수 합계 (Q~S열)
            q_cell_sum = sum(ws.cell(row=r, column=17).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            r_cell_sum = sum(ws.cell(row=r, column=18).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            s_cell_sum = sum(ws.cell(row=r, column=19).value or 0 for r in range(data_start_row, data_start_row + 4) if r <= ws.max_row)
            
            safe_write_stat_cell(summary_row, 17, q_cell_sum if q_cell_sum > 0 else s_sum)
            safe_write_stat_cell(summary_row, 18, r_cell_sum if r_cell_sum > 0 else t_sum)
            safe_write_stat_cell(summary_row, 19, s_cell_sum if s_cell_sum > 0 else u_sum)
    
    # 총계 행 (23행): 7, 12, 17, 22행의 합
    total_row = 23
    if total_row <= ws.max_row:
        # 각 열별로 7, 12, 17, 22행의 합 계산 (C열부터 S열까지)
        for col in range(3, 20):  # C열(3)부터 S열(19)까지
            total_sum = 0
            for summary_row in [7, 12, 17, 22]:
                cell_value = ws.cell(row=summary_row, column=col).value
                if cell_value is not None:
                    try:
                        total_sum += float(cell_value)
                    except:
                        pass
            safe_write_stat_cell(total_row, col, total_sum if total_sum > 0 else None)

# 전체 시트 채우기
print("\n전체 시트 채우는 중...")
if '전체' in wb.sheetnames:
    ws = wb['전체']
    
    # 년도 업데이트 및 진급반 정보 채우기
    # 제목: "2025학년도 {previous_grade}학년" (이전 학년도 기준)
    # 하지만 "진급학년 학반발표" 제목의 괄호 안 학년만 학번 첫자리를 그대로 사용
    display_grade = previous_grade  # 이전 학년 (학번 첫자리 - 1) - 일반적으로 사용
    display_year = previous_year  # 2025학년도 (이전 학년도)
    current_display_year = current_year  # 2026학년도 (현재 학년도)
    
    # I4, L4, O4, R4 셀(9열, 12열, 15열, 18열, 4행)의 2022를 현재 연도로 변경
    # I4: 9열, L4: 12열, O4: 15열, R4: 18열
    columns_to_update = [9, 12, 15, 18]  # I, L, O, R 열
    for col in columns_to_update:
        cell = ws.cell(row=4, column=col)
        if cell.value:
            if isinstance(cell.value, str) and '2022' in str(cell.value):
                cell.value = str(cell.value).replace('2022', str(current_year))
            elif isinstance(cell.value, (int, float)) and cell.value == 2022:
                cell.value = current_year
    
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                cell_str = str(cell.value)
                # 진급학년 학반발표 제목 패턴: "2026학년도 진급학년 학반발표(2025학년도 3학년)" 형식
                # 괄호 안의 학년 정보만 학번 첫자리(current_grade)를 그대로 사용
                if '진급학년' in cell_str and '학반발표' in cell_str:
                    # 괄호 안의 학년 정보를 current_grade(학번 첫자리)로 변경 (정규표현식으로 숫자 학년 찾기)
                    # 패턴: (숫자학년도 공백)(숫자학년)
                    # replacement string에서 백레퍼런스와 변수를 올바르게 처리
                    # \1은 첫 번째 그룹 (숫자학년도 공백), current_grade는 학번 첫자리
                    replacement = f'(\\g<1>{current_grade}학년)'
                    cell.value = re.sub(r'\((\d+학년도\s+)(\d+)학년\)', replacement, cell_str)
                    # 2025학년도 부분도 업데이트
                    if '2025학년도' in str(cell.value):
                        cell.value = str(cell.value).replace('2025학년도', f'{display_year}학년도')
                    # 2026학년도 부분도 업데이트
                    if '2026학년도' in str(cell.value):
                        cell.value = str(cell.value).replace('2026학년도', f'{current_display_year}학년도')
                # 일반적인 2025학년도와 학년 패턴 업데이트
                elif '2025학년도' in cell_str and re.search(r'\d+학년', cell_str):
                    # 괄호 안의 학년 정보를 previous_grade로 변경
                    replacement2 = f'(\\g<1>{display_grade}학년)'
                    cell.value = re.sub(r'\((\d+학년도\s+)(\d+)학년\)', replacement2, cell_str)
                    # 2025학년도 부분도 업데이트
                    cell.value = str(cell.value).replace('2025학년도', f'{display_year}학년도')
                # 2026학년도 {current_grade}학년 -> 2025학년도 {previous_grade}학년
                if f'{current_display_year}학년도 {current_grade}학년' in cell_str:
                    cell.value = cell_str.replace(f'{current_display_year}학년도 {current_grade}학년', f'{display_year}학년도 {display_grade}학년')
                # 2022를 2026으로 (진급반 년도)
                if '2022' in cell_str and '진급반' in cell_str:
                    cell.value = cell_str.replace('2022', str(current_display_year))
                # 2022를 2026으로 (일반)
                if '2022' in cell_str and '진급반' not in cell_str:
                    cell.value = cell_str.replace('2022', str(current_display_year))
                # ○ - 1을 현재 학년 - 1로 (현재 학년 기준, 학번 첫자리)
                if '○ - 1' in cell_str:
                    cell.value = cell_str.replace('○ - 1', f'{display_grade} - 1')
                if '○ - 2' in cell_str:
                    cell.value = cell_str.replace('○ - 2', f'{display_grade} - 2')
                if '○ - 3' in cell_str:
                    cell.value = cell_str.replace('○ - 3', f'{display_grade} - 3')
                if '○ - 4' in cell_str:
                    cell.value = cell_str.replace('○ - 4', f'{display_grade} - 4')
                # "2 - 1" 같은 패턴을 "3 - 1"로 변경 (현재 학년으로)
                if re.search(r'\d+\s*-\s*\d+', cell_str):
                    cell.value = re.sub(r'(\d+)\s*-\s*(\d+)', f'{display_grade} - \\2', cell_str)
                # 2024학년도를 이전 학년도로
                if '2024학년도' in cell_str:
                    cell.value = cell_str.replace('2024학년도', f'{display_year}학년도')
                # 괄호 안의 2024를 이전 학년도로
                if '(2024)' in cell_str:
                    cell.value = cell_str.replace('(2024)', f'({display_year})')
                # 학년 정보 업데이트 (이전 학년 기준)
                if '( 4 )' in cell_str or '(4)' in cell_str:
                    cell.value = cell_str.replace('( 4 )', f'({display_grade})').replace('(4)', f'({display_grade})')
    
    # 진급반 정보 채우기
    # 전체 시트 구조: 1~4반 전체 자료가 한 시트에 입력되어야 함
    # 학번을 기준으로 배정반 찾기
    student_to_target_class = {}
    student_to_prev_class = {}
    for prev_class in [1, 2, 3, 4]:
        for target_class in ['A', 'B', 'C', 'D']:
            for student in previous_class_data[prev_class][target_class]['male'] + previous_class_data[prev_class][target_class]['female']:
                student_to_target_class[student['학번']] = target_class
                student_to_prev_class[student['학번']] = prev_class
    
    # 전체 시트 구조: G6부터 시작
    # GHI: 1반 (번호, 이름, 진급반)
    # JKL: 2반 (번호, 이름, 진급반)
    # MNO: 3반 (번호, 이름, 진급반)
    # PQR: 4반 (번호, 이름, 진급반)
    
    # 각 이전 학반별 학생 목록 준비
    prev_class_students = {}
    for prev_class in [1, 2, 3, 4]:
        students = []
        for target_class in ['A', 'B', 'C', 'D']:
            students.extend(previous_class_data[prev_class][target_class]['male'])
            students.extend(previous_class_data[prev_class][target_class]['female'])
        # 학번순 정렬
        students.sort(key=lambda x: int(x['학번']) if x['학번'].isdigit() else 9999)
        prev_class_students[prev_class] = students
    
    # 데이터 시작 행: G6 (6행, 7번 컬럼)
    data_start_row = 6
    data_start_col = 7  # G열
    
    # 컬럼 매핑: 각 반별로 3개 컬럼씩
    # 1반: G(7), H(8), I(9) - 번호, 이름, 진급반
    # 2반: J(10), K(11), L(12) - 번호, 이름, 진급반
    # 3반: M(13), N(14), O(15) - 번호, 이름, 진급반
    # 4반: P(16), Q(17), R(18) - 번호, 이름, 진급반
    class_columns = {
        1: {'번호': 7, '이름': 8, '진급반': 9},
        2: {'번호': 10, '이름': 11, '진급반': 12},
        3: {'번호': 13, '이름': 14, '진급반': 15},
        4: {'번호': 16, '이름': 17, '진급반': 18}
    }
    
    # 병합된 셀 범위 확인
    merged_ranges = list(ws.merged_cells.ranges)
    
    def get_merged_cell_coord(row, col):
        """병합된 셀의 경우 첫 번째 셀의 좌표 반환"""
        for merged_range in merged_ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                return (merged_range.min_row, merged_range.min_col)
        return (row, col)
    
    def safe_write_cell(row, col, value):
        """병합된 셀을 고려하여 안전하게 셀에 쓰기"""
        merge_row, merge_col = get_merged_cell_coord(row, col)
        if merge_row == row and merge_col == col:  # 첫 번째 셀인 경우만 쓰기
            ws.cell(row=row, column=col).value = value
    
    # 학번을 기준으로 배정반 찾기
    student_to_target_class = {}
    for prev_class in [1, 2, 3, 4]:
        for target_class in ['A', 'B', 'C', 'D']:
            for student in previous_class_data[prev_class][target_class]['male'] + previous_class_data[prev_class][target_class]['female']:
                student_to_target_class[student['학번']] = target_class
    
    print(f"  데이터 시작 위치: 행 {data_start_row}, 컬럼 G(7)")
    print(f"  1반 학생 수: {len(prev_class_students[1])}, 2반: {len(prev_class_students[2])}, 3반: {len(prev_class_students[3])}, 4반: {len(prev_class_students[4])}")
    
    # 전체 탭 헤더에 신 학반과 현학반 정보 채우기
    # 신 학반: 4의 A, 4의 B, 4의 C, 4의 D (current_grade + 1)
    # 현학반: 3의 1, 3의 2, 3의 3, 3의 4 (current_grade)
    new_grade = current_grade + 1  # 4학년 (진급 후)
    
    # 헤더 행 찾기 및 신 학반/현학반 정보 채우기
    # 전체 시트를 스캔하여 신 학반과 현학반 정보 찾기
    for row_idx in range(1, min(data_start_row, 10)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and isinstance(cell_value, str):
                cell_str = str(cell_value)
                # 신 학반 정보: "○의 A", "3의 A" 등을 "4의 A"로 변경
                if '의' in cell_str and ('A' in cell_str or 'B' in cell_str or 'C' in cell_str or 'D' in cell_str):
                    if '○의' in cell_str or f'{current_grade}의' in cell_str:
                        for target_class in ['A', 'B', 'C', 'D']:
                            if target_class in cell_str:
                                new_value = f'{new_grade}의 {target_class}'
                                safe_write_cell(row_idx, col_idx, new_value)
                                break
                # 현학반 정보: "2의 1", "○의 1" 등을 "3의 1"로 변경
                # 정규표현식으로 "숫자의 숫자" 또는 "○의 숫자" 패턴 찾기
                if re.search(r'[0-9○]\s*의\s*[1-4]', cell_str):
                    # "2의 1", "○의 1" 등을 "3의 1" 형식으로 변경
                    for prev_class in [1, 2, 3, 4]:
                        if re.search(r'[0-9○]\s*의\s*' + str(prev_class), cell_str):
                            new_value = f'{current_grade}의 {prev_class}'
                            safe_write_cell(row_idx, col_idx, new_value)
                            break
    
    # 각 반별로 데이터 채우기
    # 1~20번: GHI, JKL, MNO, PQR (7~18열)
    # 40번 이후: 몇 칸 내려와서 24열부터 시작 (X열부터)
    max_first_section = 20  # 첫 번째 섹션에 들어갈 최대 학생 수 (1~20번)
    second_section_start_number = 40  # 두 번째 섹션 시작 번호 (40번 이후)
    
    for prev_class in [1, 2, 3, 4]:
        students = prev_class_students[prev_class]
        cols = class_columns[prev_class]
        
        for idx, student in enumerate(students):
            # 번호 - 학번의 뒤 두 자리
            student_id = str(student.get('학번', ''))
            if len(student_id) >= 2:
                number_str = student_id[-2:]  # 뒤 두 자리
                try:
                    number = int(number_str)
                except:
                    continue  # 번호가 없으면 입력 안 함
            else:
                continue  # 번호가 없으면 입력 안 함
            
            if number <= max_first_section:
                # 1~20번: 기존 위치 (GHI, JKL, MNO, PQR)
                row = data_start_row + (number - 1)
                if row > ws.max_row:
                    break
                
                safe_write_cell(row, cols['번호'], number)
                safe_write_cell(row, cols['이름'], student.get('이름', ''))
                target_class_val = student_to_target_class.get(student['학번'], '')
                if target_class_val:
                    safe_write_cell(row, cols['진급반'], target_class_val)
            elif number >= second_section_start_number:
                # 40번 이후: 24행부터 시작, 각 반별로 G열부터 시작
                # 1반 41번: G24(7열, 24행)부터 시작
                # 2반 41번: J24(10열, 24행)부터 시작
                # 3반 41번: M24(13열, 24행)부터 시작
                # 4반 41번: P24(16열, 24행)부터 시작
                second_section_start_row = 24  # 24행부터 시작
                second_section_start_col = 7 + (prev_class - 1) * 3  # 1반: 7열(G), 2반: 10열(J), 3반: 13열(M), 4반: 16열(P)
                # 41번이 24행에 오도록: 24 + (number - 41)
                row = second_section_start_row + (number - 41)
                if row > ws.max_row:
                    break
                
                safe_write_cell(row, second_section_start_col, number)  # 번호
                safe_write_cell(row, second_section_start_col + 1, student.get('이름', ''))  # 이름
                target_class_val = student_to_target_class.get(student['학번'], '')
                if target_class_val:
                    safe_write_cell(row, second_section_start_col + 2, target_class_val)  # 진급반

# ------------------------------------------------------------
# 반편성_완료.xlsx의 '반편성 배정표' 시트를 출력서식_완료.xlsx에도 포함
# ------------------------------------------------------------
def copy_sheet_between_workbooks(src_wb, src_sheet_name, dst_wb, dst_sheet_name=None):
    """다른 워크북의 시트를 dst_wb로 복사(값/스타일/병합/행높이/열너비)."""
    if src_sheet_name not in src_wb.sheetnames:
        print(f"  경고: 원본에 '{src_sheet_name}' 시트가 없습니다.")
        return

    src_ws = src_wb[src_sheet_name]
    dst_name = dst_sheet_name or src_sheet_name

    # 같은 이름 시트가 이미 있으면 삭제 후 재생성
    if dst_name in dst_wb.sheetnames:
        del dst_wb[dst_name]
    dst_ws = dst_wb.create_sheet(title=dst_name)

    # 열 너비
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    # 행 높이
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height

    # 셀 값/스타일 복사
    for row in src_ws.iter_rows():
        for cell in row:
            # MergedCell은 col_idx가 없을 수 있어 column 사용
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 병합 셀 복사
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    print(f"  ✓ '{src_sheet_name}' 시트를 '{dst_name}'로 복사 완료")

try:
    completed_wb = load_workbook("반편성_완료.xlsx")
    copy_sheet_between_workbooks(completed_wb, "반편성 배정표", wb, "반편성 배정표")
except Exception as e:
    print(f"  경고: 반편성 배정표 시트 복사 중 오류: {e}")

# 파일 저장
output_file = "출력서식_완료.xlsx"
try:
    wb.save(output_file)
    print(f"\n✓ 파일이 생성되었습니다: {output_file}")
except PermissionError:
    import time
    output_file = f"출력서식_완료_{int(time.time())}.xlsx"
    wb.save(output_file)
    print(f"\n✓ 파일이 생성되었습니다: {output_file}")

print("\n완료!")
